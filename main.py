"""
Fleet DQP Audit App
GroundProbe · Multi-user · SharePoint / Local / Railway

Storage modes (set via env var DQP_STORAGE):
  local      — reads/writes Excel from DQP_EXCEL_PATH (default)
  sharepoint — reads/writes via Microsoft Graph API

SharePoint env vars (only needed for sharepoint mode):
  SP_TENANT_ID, SP_CLIENT_ID, SP_CLIENT_SECRET, SP_SITE_ID, SP_FILE_PATH
"""

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime, date
import os, re, io, time

# PDF generation
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

try:
    from msal import ConfidentialClientApplication
    import requests as _requests
    MSAL_AVAILABLE = True
except ImportError:
    MSAL_AVAILABLE = False

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
STORAGE_MODE = os.environ.get("DQP_STORAGE", "local")
EXCEL_PATH   = os.environ.get("DQP_EXCEL_PATH", "Fleet_DQP_Master.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# BUSINESS UNIT MAPPING  (Site → BU)
# Add new sites here as the fleet expands globally
# ─────────────────────────────────────────────────────────────────────────────
SITE_BU = {
    # ── GPNA — North America (USA, Canada, Mexico) ────────────────────────────
    "Arcelormittal":        "GPNA",
    "Asarco Mission":       "GPNA",
    "Asarco Ray":           "GPNA",
    "Bingham":              "GPNA",
    "Bloom Lake":           "GPNA",
    "Buena Vista del Cobre":"GPNA",
    "Canadian Malartic":    "GPNA",
    "Chino":                "GPNA",
    "Copper Mountain":      "GPNA",
    "Copper mountain":      "GPNA",
    "Diavik":               "GPNA",
    "Ekati":                "GPNA",
    "Elkview":              "GPNA",
    "Gahcho Kue":           "GPNA",
    "Gibraltar":            "GPNA",
    "Grande Cache":         "GPNA",
    "HVC":                  "GPNA",
    "La Caridad":           "GPNA",
    "Libby Dam":            "GPNA",
    "Magino":               "GPNA",
    "Meadowbank":           "GPNA",
    "Meliadine":            "GPNA",
    "Morenci":              "GPNA",
    "Mount Milligan":       "GPNA",
    "Mulatos":              "GPNA",
    "Murray Pit":           "GPNA",
    "Origin Mine":          "GPNA",
    "Peñasquito":           "GPNA",
    "Pinos Altos":          "GPNA",
    "Pinto Valley":         "GPNA",
    "Porcupine":            "GPNA",
    "Questa":               "GPNA",
    "Red Chris":            "GPNA",
    "Sierrita":             "GPNA",
    "Thiess":               "GPNA",
    "Thompson Creek":       "GPNA",
    "Torex":                "GPNA",
    "Tyrone":               "GPNA",
    "Victoria Eagle Gold":  "GPNA",
    # ── GPCL — Chile ─────────────────────────────────────────────────────────
    "Alumbrera":            "GPCL",
    "Andina":               "GPCL",
    "Antucoya":             "GPCL",
    "CMP":                  "GPCL",
    "Candelaria":           "GPCL",
    "Caserones":            "GPCL",
    "Chuquicamata":         "GPCL",
    "Collahuasi":           "GPCL",
    "DMH":                  "GPCL",
    "El Salvador":          "GPCL",
    "El Soldado":           "GPCL",
    "Gaby":                 "GPCL",
    "Lomas Bayas":          "GPCL",
    "Mantos Blancos":       "GPCL",
    "Mantoverde":           "GPCL",
    "Minera San Cristobal": "GPCL",
    "Quebrada Blanca":      "GPCL",
    "Radomiro Tomic":       "GPCL",
    "Salares Norte":        "GPCL",
    "Sierra Gorda":         "GPCL",
    "Zaldivar":             "GPCL",
    # ── GPPE — Peru ──────────────────────────────────────────────────────────
    "Alpamarca":            "GPPE",
    "Antamina":             "GPPE",
    "Antapaccay":           "GPPE",
    "Atacocha":             "GPPE",
    "Cerro Corona":         "GPPE",
    "Cerro Verde":          "GPPE",
    "Cerro verde":          "GPPE",
    "Chinalco Toromocho":   "GPPE",
    "Coimolache":           "GPPE",
    "El Brocal":            "GPPE",
    "Iscaycruz":            "GPPE",
    "La Arena":             "GPPE",
    "Marcobre":             "GPPE",
    "Quellaveco":           "GPPE",
    "Shahuindo":            "GPPE",
    "Summa Gold":           "GPPE",
    "Volcan":               "GPPE",
    "Yanacocha":            "GPPE",
    # ── GPCO — Colombia ──────────────────────────────────────────────────────
    "EPM":                  "GPCO",
    # ── GPCN — China ─────────────────────────────────────────────────────────
    "Hami":                 "GPCN",
    "Hualian":              "GPCN",
    "Jiama":                "GPCN",
    "LCRGE":                "GPCN",
    "Luishia":              "GPCN",
    "Pacific Mining":       "GPCN",
    "Zhungeer":             "GPCN",
    "Zijin Mine":           "GPCN",
    # ── GPID — Indonesia / India ──────────────────────────────────────────────
    "Adani GP3":            "GPID",
    "Adani PEKB":           "GPID",
    "Dipka":                "GPID",
    "Gevra":                "GPID",
    "Kusmunda":             "GPID",
    "Ostapal":              "GPID",
    "Rampura Agucha":       "GPID",
    "Sesa Goa":             "GPID",
    "Sujyoti MCL":          "GPID",
    # ── GPPT — Southeast Asia / Pacific ──────────────────────────────────────
    "AGM":                  "GPPT",
    "AMM MIFA":             "GPPT",
    "AMM SBS":              "GPPT",
    "Adaro":                "GPPT",
    "Agincourt":            "GPPT",
    "Arutmin Kintap":       "GPPT",
    "BAU":                  "GPPT",
    "BIB":                  "GPPT",
    "BSS Tabang":           "GPPT",
    "BSSR":                 "GPPT",
    "Balangan Coal":        "GPPT",
    "Bayan Tabang":         "GPPT",
    "Berau Coal":           "GPPT",
    "Bukit Asam":           "GPPT",
    "CK BIB":               "GPPT",
    "CK BMB":               "GPPT",
    "CK MHU":               "GPPT",
    "Carmen Copper":        "GPPT",
    "EGAT":                 "GPPT",
    "IMK":                  "GPPT",
    "JRBM":                 "GPPT",
    "KIM BBU":              "GPPT",
    "KPC":                  "GPPT",
    "KPP Indexim":          "GPPT",
    "Kideco":               "GPPT",
    "MAS":                  "GPPT",
    "Mitrabara Adiperdana": "GPPT",
    "Oyu Tolgoi":           "GPPT",
    "PAMA BTSJ":            "GPPT",
    "PAMA Baya":            "GPPT",
    "PAMA KPCS":            "GPPT",
    "PAMA KPCT":            "GPPT",
    "PAMA Kideco":          "GPPT",
    "PAMA MTBU":            "GPPT",
    "PPA BIB":              "GPPT",
    "PPA Bukit Asam":       "GPPT",
    "PTFI":                 "GPPT",
    "Perkasa Inakakerta":   "GPPT",
    "Petrosea":             "GPPT",
    "SBS":                  "GPPT",
    "Semirara":             "GPPT",
    "Usukh Zoos":           "GPPT",
    "Wahana Baratama Mining":"GPPT",
    # ── GPSA — South Africa / Africa & Middle East ────────────────────────────
    "Agbaou Gold Mine":                  "GPSA",
    "Amandabelt Mine":                   "GPSA",
    "Armenia ZCMC Mine":                 "GPSA",
    "Asanko Gold Mine":                  "GPSA",
    "Bibiani Gold Mine (Mensin)":        "GPSA",
    "Chirano Gold Mine":                 "GPSA",
    "Damang Gold Mine":                  "GPSA",
    "Damtshaa Mine":                     "GPSA",
    "Gamsberg Mine":                     "GPSA",
    "Geita Gold Mine":                   "GPSA",
    "Ghana Manganese Mine":              "GPSA",
    "Hounde Gold Mine":                  "GPSA",
    "ITY Gold Mine":                     "GPSA",
    "Iduapriem Mine":                    "GPSA",
    "Ivrindi":                           "GPSA",
    "Jwaneng Mne":                       "GPSA",
    "Kamoto Copper - KOV Mine":          "GPSA",
    "Kao Mine":                          "GPSA",
    "Karowe Lucara Mine":                "GPSA",
    "Kimberly":                          "GPSA",
    "Kolomela Mine":                     "GPSA",
    "Kudumane Manganese Resources mine": "GPSA",
    "Lafigue":                           "GPSA",
    "Lapseki":                           "GPSA",
    "Letlhakane Mine":                   "GPSA",
    "Mana Gold Mine":                    "GPSA",
    "Moatize":                           "GPSA",
    "Mogalakwena Mine":                  "GPSA",
    "Mokala Mine":                       "GPSA",
    "Muruntau Mine Uzberkistan":         "GPSA",
    "Mutanda Mine":                      "GPSA",
    "Nchanga Mine":                      "GPSA",
    "Nkomati Joint Venture":             "GPSA",
    "Orapa Mine":                        "GPSA",
    "Rossing Mine":                      "GPSA",
    "Sabodala Massawa Mine":             "GPSA",
    "Sandfire Motheo":                   "GPSA",
    "Siguiri Gold Mine":                 "GPSA",
    "Sishen Mine":                       "GPSA",
    "Stevin Rock":                       "GPSA",
    "Tarkwa Gold Mine":                  "GPSA",
    "Tenke Mine":                        "GPSA",
    "Terrafame Aki":                     "GPSA",
    "Thabazimbi Mine":                   "GPSA",
    "Tshipi Borwa":                      "GPSA",
    "UMK":                               "GPSA",
    "YK Mine Turkey":                    "GPSA",
    # ── GPNA additions (previously Other) ────────────────────────────────────
    "Piedras Verdes":        "GPNA",
    "Buena Vista Del Cobre": "GPNA",
    "Hvc":                   "GPNA",

}
SP_TENANT_ID = os.environ.get("SP_TENANT_ID", "")
SP_CLIENT_ID = os.environ.get("SP_CLIENT_ID", "")
SP_CLIENT_SEC= os.environ.get("SP_CLIENT_SECRET", "")
SP_SITE_ID   = os.environ.get("SP_SITE_ID", "")
SP_FILE_PATH = os.environ.get("SP_FILE_PATH", "/DQP/Fleet_DQP_Master.xlsx")

# ── Auto-bootstrap: if volume path doesn't exist, seed from repo ─────────────
# This runs once on first deploy. After that the volume persists across redeploys.
if STORAGE_MODE == "local" and not os.path.exists(EXCEL_PATH):
    import glob, shutil
    # Look for any Excel seed file in the app directory
    seeds = (
        glob.glob("Fleet_DQP_Master*.xlsx") +
        glob.glob("Fleet_DQP_Master*.xlsm") +
        glob.glob("/app/Fleet_DQP_Master*.xlsx")
    )
    if seeds:
        os.makedirs(os.path.dirname(EXCEL_PATH) or ".", exist_ok=True)
        shutil.copy(seeds[0], EXCEL_PATH)
    else:
        # No seed found — will show a clear error
        pass

# ─────────────────────────────────────────────────────────────────────────────
# DQP MASTER DICTIONARY
# ─────────────────────────────────────────────────────────────────────────────
DQP_MASTER = {
    "System Health | Data Availability": {"common": [
        "🟢 Live SSR data available at PMP / Checked daily",
        "⚠️ Intermittent data available",
        "🔴 No live SSR data available",
    ]},
    "System Health | SSR Type & Scan Mode": {
        "3D": ["🟢 Range<1400m & Short Range","🟢 Range<2800m & Long Range (RPN)","⚠️ Range>2800m but Short","🔴 Wrong Config"],
        "2D": ["🟢 FX < 3.5 Km & Correct","⚠️ FX > 3.5 Km","🟢 OMNI < 5.6 Km & Correct","⚠️ OMNI > 5.6 Km","🔴 Wrong Config"],
    },
    "System Health | Signal Strength": {
        "3D": ["🟢 Amplitude consistent across wall / No low periods","⚠️ Gradual reduction / Eventual dips","🔴 Similar strength Sky vs Rock / Blue Amplitude"],
        "2D": ["🟢 Amplitude consistent across wall / No low periods","⚠️ Gradual reduction / Eventual dips","🔴 Similar strength Sky vs Rock / Blue Amplitude"],
    },
    "System Health | Return Signal": {
        "3D": ["🟢 The ΔCoherence image for the rock face is generally clear (white)","⚠️ The ΔCoherence image shows large areas of dark colours","🔴 The ΔCoherence image cannot be improved by filtering"],
        "2D": ["🟢 Coherence clear / No dark areas","⚠️ Dark areas (filtered out)","🔴 The Coherence image cannot be improved by filtering using the time slider"],
    },
    "System Health | Wall %": {
        "3D": ["⚪ N/A (Applies to 2D)"],
        "2D": ["🟢 No alarms / Minimal floating pixels","⚠️ Disturbing amount of floating pixels","🔴 Wall % > 30% / False alarms triggering"],
    },
    "Scan Area | Selection / Width": {
        "3D": ["🟢 < 1/8th scan unnecessary","⚠️ ~ 1/4 scan unnecessary","🔴 > 1/4 scan unnecessary","🔴 Area of interest missing from scan"],
        "2D": ["🟢 Areas minimised (Optimal)","⚠️ Only 60deg or 120deg used (not HD mode)","🔴 Full Scan 60/120deg Used (Unnecessary)","🔴 Area of interest missing from scan"],
    },
    "Scan Area | Elevation": {
        "3D": ["⚪ N/A (Applies to 2D)"],
        "2D": ["🟢 60deg view used well / STC avoided","⚠️ Too much pit bottom or sky","🔴 Pointing wrong / Missing area"],
    },
    "Scan Area | Levelling": {
        "3D": ["⚪ N/A (Applies to 2D)"],
        "2D": ["🟢 Properly levelled","🔴 Not properly levelled"],
    },
    "Scan Area | Coverage / Angle": {
        "3D": ["🟢 Good incidence (>75% mag)","⚠️ Poor incidence (50% mag)","🔴 Bad incidence (30% mag)"],
        "2D": ["🟢 Plan view matches front view","⚠️ Geo shows missing areas (limits of 60deg)","🔴 Geo shows large misalignment"],
    },
    "Photographs | Camera Alignment": {"common": [
        "🟢 Aligned (<= 2px error)","⚠️ Misaligned (3-4px error) / Shifting photos","🔴 Misaligned (>4px / Shift)",
    ]},
    "Photographs | Photo Quality": {"common": [
        "🟢 Clear and usable","⚠️ Hard to interpret / Dirty Lens / Foggy","🔴 Unusable",
    ]},
    "Photographs | Geopositioning": {
        "3D": ["⚪ N/A (Applies to 2D)"],
        "2D": ["🟢 Accurate / Layers match","⚠️ Inaccurate / Old image","🔴 Not made"],
    },
    "Data Visualization | DTM & Geopositioning": {
        "3D": ["⚪ N/A (Applies to 2D)"],
        "2D": ["🟢 DTM and Geopositioning applied correctly","⚠️ Only DTM or Geopositioning applied","🔴 DTM and Geopositioning not applied"],
    },
    "Masks | Sky & Short Range": {
        "3D": ["🟢 Only sky/short range (<30m) masked out","⚠️ Sky/short range NOT automatically masked","🔴 Mask amplitude too high / Masking Area of Interest"],
        "2D": ["🟢 Only sky/short range targets masked out","⚠️ Range wrapped targets triggering false alarms","🔴 Sky mask amplitude too low / Excess pixels","🔴 Mask amplitude too high / Masking Area of Interest"],
    },
    "Masks | Manual Mask": {"common": [
        "🟢 Manual masks used correctly","⚠️ Loose masking of equipment/veg","🔴 Missing manual masks","🔴 Area of interest masked out",
    ]},
    "Masks | EDM": {"common": [
        "🟢 Applied correctly OR Not Required","⚠️ Partially applied","🔴 Not applied (Risk) OR Whole area masked",
    ]},
    "Alarming | Watchdog": {"common": [
        "🟢 Critical Monitoring / Active / Tested",
        "⚠️ Watchdog active but sounding alarms",
        "🟢 Not required (Customer Decision)",
        "🔴 Watchdog Disabled / Setup Mode",
    ]},
    "Alarming | Settings": {"common": [
        "🟢 Alarms set with thresholds/notifications","🟢 No alarms (Customer TARP Decision)","🔴 No alarms / Thresholds without notifications",
    ]},
    "Alarming | Tracking": {"common": [
        "🟢 Alarm set for fast movement","🟢 Configured / Not Necessary","🔴 Not set (Ambiguity risk)",
    ]},
    "Alarming | Coherence Area": {
        "3D": ["🟢 ΔCoherence (0-0.3) White/Light Blue","🔴 ΔCoherence Blue/Dark Blue"],
        "2D": ["🟢 Alarms on areas of high coherence","🔴 Alarms on areas of consistent low coherence"],
    },
    "Alarming | Incidence Angle": {
        "3D": ["🟢 Vector loss considered","🔴 Vector loss NOT considered"],
        "2D": ["🟢 Vector loss considered in alarm thresholds","⚠️ 70% magnitude - no alarm adjustment","🔴 >30% vector loss / No alarm adjustment"],
    },
    "Atmospheric | Algorithm": {"common": [
        "🟢 Enhanced Deformation algorithm is selected","⚠️ Standard Deformation algorithm is selected","🔴 Predictive Deformation algorithm is selected",
    ]},
    "Atmospheric | Source": {
        "3D": [
            "🟢 SRAs / DSRAs Selected",
            "⚠️ Weather Station (WS) Selected / SSR < 600m from wall",
            "🔴 WS selected / SSR > 600m from wall",
            "🔴 No atmospheric correction applied"
        ],
        "2D": [
            "🟢 Precision Atmospherics (PA) selected",
            "⚠️ DSRAs selected",
            "⚠️ WS selected / SSR < 600m from wall / SRA selected",
            "🔴 WS selected / SSR > 600m from wall",
            "🔴 No atmospheric correction applied"
        ],
    },
    "Atmospheric | Created SRAs": {
        "3D": ["🟢 2+ SRAs independent domains","⚠️ SRAs on same domain or similar range","🔴 Only 1 SRA created","🔴 None"],
        "2D": ["🟢 2+ DSRAs/SRAs independent domains & ranges","⚠️ DSRAs/SRAs on same domain or similar range","🔴 Only 1 DSRA/SRA created","🔴 None / Poor quality"],
    },
    "Atmospheric | SRA Spread Graph": {"common": [
        "🟢 Horizontal trend, no steps","⚠️ The data exhibits a stepwise pattern or a regressive line","🔴 The plot shows an increasing deformation spread",
    ]},
    "Atmospheric | Stable Reference Pixels": {
        "3D": ["🟢 All SRP White (DCoherence <0.1)","⚠️ Some SRP in noisy areas (<1/3)","🔴 SRP Dark (DCoh > 0.1)"],
        "2D": ["🟢 All SRP White (Coherence >0.98)","⚠️ Some SRP low coherence (0.90-0.98)","🔴 SRP Dark (Coh < 0.90)"],
    },
    "Atmospheric correction | Graph (1 Day)": {"common": [
        "🟢 Plots follow same trend","⚠️ The plots exhibit parallel trends, despite minor offsets","🔴 Plots diverge",
    ]},
    "Atmospheric | Rejected SRA %": {"common": [
        "🟢 < 10% Rejected","⚠️ 10-30% Rejected","🔴 > 30% Rejected",
    ]},
    "Atmospheric | WS Graph (2 Days)": {
        "3D": ["⚪ N/A (SRAs used)","🟢 Sinusoidal & Horizontal Trend","⚠️ Sinusoidal & Trend (Temp related)","🔴 Sinusoidal & Trend (Stable Weather)"],
        "2D": ["⚪ N/A (SRAs/PA used)","🟢 Sinusoidal & Horizontal Trend","⚠️ Sinusoidal & Trend (Temp related)","🔴 Sinusoidal & Trend (Stable Weather)"],
    },
    "Atmospheric | Created PA": {
        "3D": ["⚪ N/A (Applies to 2D PA)"],
        "2D": [
            "⚪ N/A (PA not used)",
            "🟢 Seed Mask at centre / Good coherence / No movement / Several pixels",
            "🟢 Bulk Areas on independent domains / Different ranges & azimuths / Good coherence / No movement",
            "⚠️ Bulk Areas all at similar range",
            "🔴 Bulk Areas on moving / unstable areas"
        ],
    },
    "Atmospheric | PA Refractivity Graph": {
        "3D": ["⚪ N/A (Applies to 2D PA)"],
        "2D": [
            "⚪ N/A (PA not used)",
            "🟢 WS+PA plots similar trend / WS+SRA plots upside-down reflection (WS mode)",
            "⚠️ Step due to atmospheric event / General trend still similar (WS mode)",
            "🔴 WS and PA plots diverging continuously (WS mode)",
            "🟢 WS, SRA and PA plots all follow similar trend (SRA/DSRA mode)",
            "⚠️ Step due to atmospheric event / General trend still similar (SRA/DSRA mode)",
            "🔴 WS, SRA and PA plots diverging from each other (SRA/DSRA mode)"
        ],
    },
    "Atmospheric | PA Deformation Map": {
        "3D": ["⚪ N/A (Applies to 2D PA)"],
        "2D": [
            "⚪ N/A (PA not used)",
            "🟢 Little or no contamination / All deformations make geotechnical sense",
            "⚠️ Few regions with minor contamination / Most deformation makes sense",
            "⚠️ Displacement on Bulk Areas or Seed Mask / Reasonable in short-term windows",
            "🔴 Deformation impossible kinematically / Does not make geotechnical sense"
        ],
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# PMP HEALTH CHECKPOINTS (separate from DQP score)
# ─────────────────────────────────────────────────────────────────────────────
PMP_HEALTH = {
    "PMP Health | Last Check": {"common": [
        "🟢 < 30 days ago",
        "⚠️ 30–90 days ago",
        "🔴 > 90 days ago / Unknown",
    ]},
    "PMP Health | MonitorIQ Version": {"common": [
        "🟢 Latest version installed",
        "⚠️ 1 version behind",
        "🔴 2+ versions behind / Unknown",
    ]},
    "PMP Health | Hardware Compliance": {
        # XT: 16GB RAM rec / 8GB min, 500GB SSD, GPU DirectX11+4GB VRAM
        "3D": [
            "🟢 Meets recommended (i5-13600K+, 16GB RAM, 500GB SSD, NVIDIA 1060+)",
            "⚠️ Meets minimum (i5-12500+, 8GB RAM, 500GB SSD, DirectX11 GPU)",
            "🔴 Below minimum requirements",
        ],
        # FX: 32GB RAM rec / 16GB min, 4TB SSD
        "FX": [
            "🟢 Meets recommended (i7-13700K+, 32GB RAM, 4TB SSD, NVIDIA 1060+)",
            "⚠️ Meets minimum (i5-12500+, 16GB RAM, 4TB SSD, DirectX11 GPU)",
            "🔴 Below minimum requirements",
        ],
        # SOM/OMNI: 32GB RAM rec / 16GB min, 4TB SSD
        "SOM": [
            "🟢 Meets recommended (i7-13700K+, 32GB RAM, 4TB SSD, NVIDIA 1060+)",
            "⚠️ Meets minimum (i5-12500+, 16GB RAM, 4TB SSD, DirectX11 GPU)",
            "🔴 Below minimum requirements",
        ],
        # SAR-X: 32GB RAM rec / 16GB min, 2TB SSD
        "SAR": [
            "🟢 Meets recommended (i7-13700K+, 32GB RAM, 2TB SSD, NVIDIA 1060+)",
            "⚠️ Meets minimum (i5-12500+, 16GB RAM, 2TB SSD, DirectX11 GPU)",
            "🔴 Below minimum requirements",
        ],
    },
    "PMP Health | PMP Responsiveness": {"common": [
        "🟢 < 1 second plot delay",
        "⚠️ 1–3 seconds delay",
        "🔴 > 3 seconds / Unresponsive",
    ]},
    "PMP Health | SSD Space Available": {
        # XT: 500GB total
        "3D": [
            "🟢 > 100 GB free (of 500 GB)",
            "⚠️ 50–100 GB free",
            "🔴 < 50 GB free / Critical",
        ],
        # FX: 4TB total
        "FX": [
            "🟢 > 800 GB free (of 4 TB)",
            "⚠️ 400–800 GB free",
            "🔴 < 400 GB free / Critical",
        ],
        # SOM: 4TB total
        "SOM": [
            "🟢 > 800 GB free (of 4 TB)",
            "⚠️ 400–800 GB free",
            "🔴 < 400 GB free / Critical",
        ],
        # SAR-X: 2TB total
        "SAR": [
            "🟢 > 400 GB free (of 2 TB)",
            "⚠️ 200–400 GB free",
            "🔴 < 200 GB free / Critical",
        ],
    },
    "PMP Health | Data Backup": {"common": [
        "🟢 Active backup within local network",
        "⚠️ Backup configured but not verified",
        "🔴 No backup / Not configured",
    ]},
    "PMP Health | Webupload Logfiles": {"common": [
        "🟢 Working correctly",
        "⚠️ Intermittent / Partial upload",
        "🔴 Disabled / Not working",
    ]},
    "PMP Health | Webupload Production Data": {"common": [
        "🟢 Wall folders uploading correctly",
        "⚠️ Partial / Intermittent upload",
        "🔴 Disabled / Not working",
    ]},
}

# PMP Health header colors (teal/slate — distinct from DQP categories)
PMP_CAT_COLOR = "#1E6B8C"

# PMP FIXES (recommended actions)
PMP_FIXES = {
    "Last Check":              "Schedule PMP maintenance visit",
    "MonitorIQ Version":       "Update MonitorIQ to latest version",
    "Hardware Compliance":     "Upgrade hardware to meet minimum requirements",
    "PMP Responsiveness":      "Check CPU/RAM usage, close background processes, consider hardware upgrade",
    "SSD Space Available":     "Free up disk space or expand storage",
    "Data Backup":             "Configure automated backup to local network",
    "Webupload Logfiles":      "Check webupload configuration and network connectivity",
    "Webupload Production Data":"Check webupload configuration for wall folders",
}

PA_DTM_MANDATORY_2D = [
    "Data Visualization | DTM & Geopositioning",
    "Atmospheric | Source",
    "Atmospheric | Created PA",
    "Atmospheric | PA Refractivity Graph",
    "Atmospheric | PA Deformation Map",
]

FIXES = {
    "Data Availability":        "Check Link & Power",
    "SSR Type":                 "Adjust Radar Parameters",
    "Signal Strength":          "Check RF cables/EC/HPA or filter low amplitudes via time slider",
    "Wall %":                   "Adjust Sky Mask Filter",
    "Return Signal":            "Check System",
    "Selection / Width":        "Optimize Scan Area",
    "Elevation":                "Adjust Elevation Limits",
    "Levelling":                "Re-level Radar Platform",
    "Coverage / Angle":         "Check Positioning / Move Radar",
    "Camera Alignment":         "Calibrate Camera",
    "Photo Quality":            "Clean Lens or change time photo",
    "Geopositioning":           "Apply Geopositioning in PMP",
    "DTM & Geopositioning":     "Apply DTM and Geopositioning in PMP",
    "Sky & Short Range":        "Check RF cables/EC/HPA or filter low amplitudes via time slider",
    "Manual Mask":              "Create or adjust Manual Masks",
    "EDM":                      "Create or adjust EDM mask properly",
    "Watchdog":                 "Enable Watchdog",
    "Settings":                 "Add Notifications to Alarms",
    "Tracking":                 "Create an alarm with Tracking threshold",
    "Coherence Area":           "Check alarm mask",
    "Incidence Angle":          "Review Alarm Vectors",
    "Algorithm":                "Select Correct Algorithm",
    "Source":                   "Define SRAs correctly",
    "Created SRAs":             "Add Independent SRAs or create more than 1",
    "SRA Spread Graph":         "Ensure SRAs are stable",
    "Stable Reference Pixels":  "Ensure SRAs are stable",
    "Graph (1 Day)":            "Ensure SRAs are stable",
    "Rejected SRA %":           "Redefine SRAs Areas",
    "WS Graph (2 Days)":        "Check Long Term Weather",
    "PA Refractivity Graph":    "Review PA Bulk Areas and Seed Mask placement",
    "Created PA":               "Redefine Seed Mask and Bulk Areas on stable non-moving zones",
    "PA Deformation Map":       "Redefine Bulk Areas on stable non-moving zones",
    "Check System":             "Check System",
}

# ─────────────────────────────────────────────────────────────────────────────
# STORAGE LAYER
# ─────────────────────────────────────────────────────────────────────────────
def _natural_key(s):
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r'([0-9]+)', str(s))]

@st.cache_data(ttl=30)
def _sp_token():
    app = ConfidentialClientApplication(SP_CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{SP_TENANT_ID}",
        client_credential=SP_CLIENT_SEC)
    return app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])["access_token"]

def _sp_download():
    token = _sp_token()
    url = f"https://graph.microsoft.com/v1.0/sites/{SP_SITE_ID}/drive/root:{SP_FILE_PATH}:/content"
    r = _requests.get(url, headers={"Authorization": f"Bearer {token}"})
    r.raise_for_status()
    return r.content

def _sp_upload(data: bytes):
    token = _sp_token()
    url = f"https://graph.microsoft.com/v1.0/sites/{SP_SITE_ID}/drive/root:{SP_FILE_PATH}:/content"
    _requests.put(url, headers={"Authorization": f"Bearer {token}", "Content-Type": "application/octet-stream"}, data=data).raise_for_status()

@st.cache_data
def load_wb_bytes() -> bytes:
    if STORAGE_MODE == "sharepoint":
        return _sp_download()
    with open(EXCEL_PATH, "rb") as f:
        return f.read()

def get_workbook():
    return load_workbook(io.BytesIO(load_wb_bytes()))

def save_workbook(wb):
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()
    if STORAGE_MODE == "sharepoint":
        _sp_upload(raw)
    else:
        with open(EXCEL_PATH, "wb") as f:
            f.write(raw)
    load_wb_bytes.clear()

# ─────────────────────────────────────────────────────────────────────────────
# DATA HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def get_audit_matrix(wb):
    ws = wb["AUDIT MATRIX"]
    headers = [cell.value for cell in ws[1]]
    rows = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    df = pd.DataFrame(rows)
    # BU: read from Excel column if it exists, fallback to SITE_BU dict
    if "BU" in df.columns:
        df["BU"] = df.apply(
            lambda r: str(r["BU"]).strip()
            if pd.notna(r.get("BU")) and str(r.get("BU","")).strip() not in ["","nan"]
            else SITE_BU.get(str(r.get("Site","")).strip(), "Other"), axis=1
        )
    elif "Site" in df.columns:
        df["BU"] = df["Site"].apply(lambda s: SITE_BU.get(str(s).strip(), "Other"))
    else:
        df["BU"] = "Other"
    return df


def ensure_bu_column(ws):
    """Ensure AUDIT MATRIX has a BU column. Returns col_index (1-based)."""
    headers = [cell.value for cell in ws[1]]
    if "BU" not in headers:
        ra_idx = headers.index("Remote Access") + 1 if "Remote Access" in headers else len(headers)
        ws.insert_cols(ra_idx + 1)
        ws.cell(1, ra_idx + 1).value = "BU"
        for row in ws.iter_rows(min_row=2):
            if not row[0].value: continue
            site_col = headers.index("Site") + 1 if "Site" in headers else 2
            site_val = str(ws.cell(row[0].row, site_col).value or "").strip()
            ws.cell(row[0].row, ra_idx + 1).value = SITE_BU.get(site_val, "Other")
        headers.insert(ra_idx, "BU")
    return headers.index("BU") + 1

def get_history(wb):
    ws = wb["HISTORY_LOG"]
    cols = ["Radar Name","Site","Audit Date","Auditor","Score Captured","Notes"]
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
    if not rows:
        return pd.DataFrame(columns=cols)
    # Pad rows to 6 cols in case Notes column doesn't exist yet
    rows = [list(r) + [None] * (6 - len(r)) for r in rows]
    df = pd.DataFrame([r[:6] for r in rows], columns=cols)
    df["Audit Date"]      = pd.to_datetime(df["Audit Date"], errors="coerce")
    df["Score Captured"]  = pd.to_numeric(df["Score Captured"], errors="coerce")
    return df

def is_3d(t): return "XT" in str(t).upper()

def get_options(key, radar_type, pa_stl="N/A"):
    data = DQP_MASTER.get(key, {})
    dim  = "3D" if is_3d(radar_type) else "2D"
    opts = data.get(dim, data.get("common", ["⚪ N/A"]))

    # PA-specific fields: only hide if radar CANNOT use PA (XT/SAR-X)
    # FX/SOM always show PA options — pa_stl='Yes' means Albert's rule applies,
    # pa_stl='No' means PA not configured but fields still visible for auditing
    if 'PA' in key:
        if not can_use_pa(radar_type):
            # XT / SAR-X → always N/A
            na_opts = [o for o in opts if '⚪' in o]
            return na_opts if na_opts else ["⚪ N/A"]
        # FX / SOM → always show full options regardless of pa_stl
    return opts

def can_use_pa(radar_type: str) -> bool:
    """Only FX and SOM (OMNI) radars can use Precision Atmospherics."""
    t = str(radar_type).upper()
    return 'FX' in t or 'SOM' in t

def compute_score(answers, radar_type, pa_stl="N/A"):
    """
    Score logic:
    - System Health 🔴 → score = 0% (data not reliable)
    - Other categories 🔴 → count in score normally
    - PA rule: ALL FX/SOM must have PA configured (no exceptions)
              XT/SAR-X → PA fields are N/A, never penalized
    """
    # Check System Health specifically
    SH_KEYS = ["System Health | Data Availability", "System Health | SSR Type & Scan Mode",
                "System Health | Signal Strength", "System Health | Return Signal",
                "System Health | Wall %",
                "Data Availability", "SSR Type & Scan Mode", "Signal Strength", "Return Signal", "Wall %"]

    sh_critical = any("🔴" in str(answers.get(k, "")) for k in SH_KEYS)

    if sh_critical:
        issues, actions = [], []
        for key, val in answers.items():
            if "🔴" in str(val) or "⚠️" in str(val):
                short   = key.split("|")[-1].strip()
                fix_key = next((k for k in FIXES if k in key), None)
                issues.append(short)
                actions.append(FIXES.get(fix_key, "Contact GroundProbe Support"))
        return 0.0, issues, actions

    green_n = sum(1 for v in answers.values() if "🟢" in str(v))
    red_n   = sum(1 for v in answers.values() if "🔴" in str(v))
    na_n    = sum(1 for v in answers.values() if "⚪" in str(v))
    total   = len(answers) - na_n

    score = green_n / total if total > 0 else 1.0

    # PA rule: ALL FX/SOM must have PA + DTM configured
    # If PA not properly set → cannot reach 100%, but score still reflects other checks
    if can_use_pa(radar_type):
        pa_src  = answers.get("Atmospheric | Source", answers.get("Source", ""))
        dtm_val = answers.get("Data Visualization | DTM & Geopositioning", answers.get("DTM & Geopositioning", ""))
        pa_map  = answers.get("Atmospheric | PA Deformation Map", answers.get("PA Deformation Map", ""))
        pa_ok   = ("Precision Atmospherics" in str(pa_src)
                   and "🟢" in str(dtm_val)
                   and "🔴" not in str(pa_map))
        if not pa_ok:
            # Cap score just below perfect — real score still shows actual state
            score = min(score, 0.99)
    issues, actions = [], []
    for key, val in answers.items():
        if "🔴" in str(val) or "⚠️" in str(val):
            short   = key.split("|")[-1].strip()
            fix_key = next((k for k in FIXES if k in key), None)
            issues.append(short)
            actions.append(FIXES.get(fix_key, "Contact GroundProbe Support"))
    return round(score, 4), issues, actions

def save_audit(radar_name, site, radar_type, auditor, audit_date, answers, score, issues, actions, notes="", pmp_answers=None, is_correction=False):
    wb       = get_workbook()
    ws_audit = wb["AUDIT MATRIX"]
    headers  = [c.value for c in ws_audit[1]]
    col_map  = {h: i+1 for i, h in enumerate(headers) if h}

    target_row = next((r[0].row for r in ws_audit.iter_rows(min_row=2) if str(r[0].value) == radar_name), None)
    if not target_row:
        return False, f"Radar {radar_name} not found"

    ws_audit.cell(target_row, col_map["Last Audit"]).value   = audit_date
    ws_audit.cell(target_row, col_map["Auditor"]).value      = auditor

    if not is_correction:
        # Normal save — increment audit count
        prev = ws_audit.cell(target_row, col_map["Total Audits"]).value or 0
        ws_audit.cell(target_row, col_map["Total Audits"]).value = int(prev) + 1

    for key, val in answers.items():
        short_key = key.split("|")[-1].strip() if "|" in key else key
        if key in col_map:
            ws_audit.cell(target_row, col_map[key]).value = val
        elif short_key in col_map:
            ws_audit.cell(target_row, col_map[short_key]).value = val
    if pmp_answers:
        for key, val in pmp_answers.items():
            short_key = key.split("|")[-1].strip() if "|" in key else key
            if key in col_map:
                ws_audit.cell(target_row, col_map[key]).value = val
            elif short_key in col_map:
                ws_audit.cell(target_row, col_map[short_key]).value = val
            else:
                new_col = max(col_map.values()) + 1
                ws_audit.cell(1, new_col).value = short_key
                ws_audit.cell(target_row, new_col).value = val
                col_map[short_key] = new_col
    if "SCORE"              in col_map: ws_audit.cell(target_row, col_map["SCORE"]).value              = score
    if "ISSUES DETECTED"    in col_map: ws_audit.cell(target_row, col_map["ISSUES DETECTED"]).value    = "\n".join(issues)
    if "RECOMMENDED ACTION" in col_map: ws_audit.cell(target_row, col_map["RECOMMENDED ACTION"]).value = "\n".join(actions)
    if "HELPER_STATUS"      in col_map: ws_audit.cell(target_row, col_map["HELPER_STATUS"]).value      = "Critical" if score==0 else ("Needs Attention" if score<1 else "Good")
    if "HELPER_SCORE"       in col_map: ws_audit.cell(target_row, col_map["HELPER_SCORE"]).value       = score

    ws_hist = wb["HISTORY_LOG"]
    if is_correction:
        # Find the last history entry for this radar and overwrite it
        last_row = None
        for r in range(2, ws_hist.max_row + 1):
            if ws_hist.cell(r, 1).value == radar_name:
                last_row = r
        if last_row:
            ws_hist.cell(last_row, 3).value = audit_date
            ws_hist.cell(last_row, 4).value = auditor
            ws_hist.cell(last_row, 5).value = score
            ws_hist.cell(last_row, 6).value = notes
        # If no history exists yet, treat as normal save
        else:
            is_correction = False

    if not is_correction:
        next_row = next((r for r in range(2, ws_hist.max_row+2) if ws_hist.cell(r,1).value is None), ws_hist.max_row+1)
        ws_hist.cell(next_row, 1).value = radar_name
        ws_hist.cell(next_row, 2).value = site
        ws_hist.cell(next_row, 3).value = audit_date
        ws_hist.cell(next_row, 4).value = auditor
        ws_hist.cell(next_row, 5).value = score
        ws_hist.cell(next_row, 6).value = notes

    save_workbook(wb)
    return True, "OK"


# ─────────────────────────────────────────────────────────────────────────────
# PDF CLIENT REPORT
# ─────────────────────────────────────────────────────────────────────────────
def generate_client_pdf(site: str, df_matrix: pd.DataFrame, auditor: str, logo_path: str = None) -> bytes:
    """Generate a professional client-facing PDF for a given site."""
    import os
    from reportlab.platypus import KeepTogether, Image as RLImage

    buf = io.BytesIO()

    GP_ORANGE  = colors.HexColor("#F78F1E")
    GP_DARK    = colors.HexColor("#393B41")
    GP_LIGHT   = colors.HexColor("#F5F6F8")
    GP_MID     = colors.HexColor("#E8EAED")
    C_GREEN    = colors.HexColor("#00B050")
    C_YELLOW   = colors.HexColor("#FFC000")
    C_RED      = colors.HexColor("#C62828")
    C_GREY     = colors.HexColor("#6B7280")
    C_LGREY    = colors.HexColor("#9CA3AF")
    WHITE      = colors.white
    C_GREEN_BG = colors.HexColor("#F0FDF4")
    C_YELLOW_BG= colors.HexColor("#FFFBEB")
    C_RED_BG   = colors.HexColor("#FEF2F2")

    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=0*mm, bottomMargin=12*mm)
    W = A4[0] - 30*mm
    story = []

    # ── Styles ────────────────────────────────────────────────────────────────
    def ps(name, **kw):
        base = dict(fontName="Helvetica", fontSize=9, textColor=GP_DARK, leading=13)
        base.update(kw)
        return ParagraphStyle(name, **base)

    s_hdr_title = ps("hdr_title", fontName="Helvetica-Bold", fontSize=18, textColor=WHITE, leading=22)
    s_hdr_sub   = ps("hdr_sub",   fontName="Helvetica",      fontSize=10, textColor=colors.HexColor("#FFE0B2"), leading=14)
    s_section   = ps("section",   fontName="Helvetica-Bold",  fontSize=11, textColor=GP_DARK, spaceBefore=10, spaceAfter=5)
    s_body      = ps("body")
    s_bold      = ps("bold",  fontName="Helvetica-Bold")
    s_small     = ps("small", fontSize=8, textColor=C_GREY, leading=11)
    s_center    = ps("center", alignment=TA_CENTER)
    s_right     = ps("right",  alignment=TA_RIGHT, fontSize=8, textColor=C_GREY)
    s_card_lbl  = ps("card_lbl", fontName="Helvetica-Bold", fontSize=8, textColor=C_LGREY, alignment=TA_CENTER, leading=10)
    s_card_val  = ps("card_val", fontName="Helvetica-Bold", fontSize=26, alignment=TA_CENTER, leading=30)
    s_radar_hdr = ps("radar_hdr", fontName="Helvetica-Bold", fontSize=9, textColor=WHITE, alignment=TA_CENTER)
    s_radar_cell= ps("radar_cell", fontSize=9, textColor=GP_DARK, alignment=TA_CENTER)
    s_radar_left= ps("radar_left", fontSize=9, textColor=GP_DARK)
    s_issue_ok  = ps("issue_ok",  fontSize=8, textColor=C_GREEN)
    s_issue_warn= ps("issue_warn", fontSize=8, textColor=colors.HexColor("#92400E"))
    s_issue_crit= ps("issue_crit", fontSize=8, textColor=C_RED)
    s_tag_green = ps("tag_g", fontName="Helvetica-Bold", fontSize=8, textColor=C_GREEN, alignment=TA_CENTER)
    s_tag_yellow= ps("tag_y", fontName="Helvetica-Bold", fontSize=8, textColor=colors.HexColor("#92400E"), alignment=TA_CENTER)
    s_tag_red   = ps("tag_r", fontName="Helvetica-Bold", fontSize=8, textColor=C_RED,   alignment=TA_CENTER)
    s_foot      = ps("foot",  fontSize=7, textColor=C_LGREY, alignment=TA_CENTER)

    today_str   = date.today().strftime("%B %d, %Y")
    today_short = date.today().strftime("%d %b %Y")

    # ── Helper: score → status ────────────────────────────────────────────────
    def score_status(s):
        if s is None: return ("Not Audited", C_GREY,   GP_LIGHT,   s_card_val)
        if s == 1.0:  return ("Good",        C_GREEN,  C_GREEN_BG, s_tag_green)
        if s >= 0.7:  return ("Needs Attention",     C_YELLOW, C_YELLOW_BG,s_tag_yellow)
        return             ("Critical",      C_RED,    C_RED_BG,   s_tag_red)

    # ── Filter site radars ────────────────────────────────────────────────────
    site_radars = df_matrix[
        df_matrix["Site"].astype(str).str.strip().str.lower() == site.strip().lower()
    ].copy()

    # DQP columns — short names (no pipe), excluding meta/calc columns
    META_COLS = {
        'Radar Name','Site','Type','Remote Access','Last Audit','Total Audits','Auditor',
        'SCORE','ISSUES DETECTED','RECOMMENDED ACTION','HELPER_SCORE','HELPER_DATE',
        'HELPER_STATUS','PA','_score'
    }
    dqp_cols = [c for c in site_radars.columns if c not in META_COLS and c and not str(c).startswith('_')]

    SH_COLS = {"Data Availability", "SSR Type & Scan Mode", "Signal Strength", "Return Signal", "Wall %"}

    def row_score(row):
        # Use HELPER_SCORE (saved score) first — consistent with what auditor saw
        saved = row.get("HELPER_SCORE")
        if saved is not None and str(saved) not in ["", "nan", "None"]:
            try:
                return float(saved)
            except (ValueError, TypeError):
                pass
        # Fallback: recalculate from checkpoints only if no saved score
        vals   = [str(row.get(c, "") or "") for c in dqp_cols]
        greens = sum(1 for v in vals if "🟢" in v)
        nas    = sum(1 for v in vals if "⚪" in v)
        total  = len(dqp_cols) - nas
        if greens + sum(1 for v in vals if "🔴" in v) + nas == 0:
            return float('nan')
        for c in SH_COLS:
            if "🔴" in str(row.get(c, "") or ""):
                return 0.0
        return greens / total if total > 0 else 1.0

    if not site_radars.empty:
        site_radars["_score"] = site_radars.apply(row_score, axis=1)
        audited = site_radars[site_radars["_score"].notna()]
        avg_score = audited["_score"].mean() if not audited.empty else None
    else:
        avg_score = None

    n_total    = len(site_radars)
    n_audited  = int(site_radars["_score"].notna().sum()) if not site_radars.empty else 0
    n_good     = int((site_radars["_score"] == 1.0).sum())   if not site_radars.empty else 0
    n_warn     = int(((site_radars["_score"] >= 0.7) & (site_radars["_score"] < 1.0)).sum()) if not site_radars.empty else 0
    n_crit     = int((site_radars["_score"].notna() & (site_radars["_score"] < 0.7)).sum())  if not site_radars.empty else 0
    n_not_aud  = n_total - n_audited

    # ── PAGE 1: COVER ─────────────────────────────────────────────────────────
    lp = logo_path or "/home/claude/gp_logo.png"
    has_logo = os.path.exists(lp)

    if has_logo:
        logo_img = RLImage(lp, width=50*mm, height=50*mm * 62/410)
        logo_cell = logo_img
    else:
        logo_cell = Paragraph("<b>GroundProbe</b>",
            ps("lc", fontName="Helvetica-Bold", fontSize=20, textColor=GP_ORANGE))

    # Dark header — orange logo looks perfect on dark background
    cover_top = Table([[
        logo_cell,
        Paragraph("Proactive Data Quality Review",
            ps("ct", fontName="Helvetica", fontSize=11, textColor=colors.HexColor("#9CA3AF"), alignment=TA_RIGHT))
    ]], colWidths=[W*0.5, W*0.5])
    cover_top.setStyle(TableStyle([
        ('BACKGROUND',   (0,0),(-1,-1), GP_DARK),
        ('VALIGN',       (0,0),(-1,-1), 'MIDDLE'),
        ('LEFTPADDING',  (0,0),(-1,-1), 12),
        ('RIGHTPADDING', (0,0),(-1,-1), 12),
        ('TOPPADDING',   (0,0),(-1,-1), 14),
        ('BOTTOMPADDING',(0,0),(-1,-1), 14),
    ]))
    story.append(cover_top)
    story.append(Spacer(1, 8*mm))

    # Site title
    story.append(Paragraph(f"<b>{site}</b>",
        ps("site_title", fontName="Helvetica-Bold", fontSize=28, textColor=GP_DARK, leading=32, spaceAfter=2)))
    story.append(Paragraph(f"Data Quality Parameters  ·  {today_str}",
        ps("site_sub", fontName="Helvetica", fontSize=11, textColor=C_GREY, spaceAfter=2)))
    story.append(Paragraph(f"Prepared by: <b>{auditor}</b>  ·  GroundProbe Advisory Services",
        ps("site_by", fontName="Helvetica", fontSize=9, textColor=C_GREY)))
    story.append(Spacer(1, 6*mm))
    story.append(HRFlowable(width=W, thickness=2, color=GP_ORANGE, spaceAfter=6*mm))

    # ── Score cards ─────────────────────────────────────────────────────────
    def make_card(label, value, val_color, bg_color=GP_LIGHT, subtitle=None):
        rows = [
            [Paragraph(label, ps("cl", fontName="Helvetica-Bold", fontSize=7.5, textColor=C_LGREY, alignment=TA_CENTER))],
            [Paragraph(str(value), ps("cv", fontName="Helvetica-Bold", fontSize=22, textColor=val_color, alignment=TA_CENTER, leading=26))],
        ]
        if subtitle:
            rows.append([Paragraph(subtitle, ps("cs", fontSize=7, textColor=C_LGREY, alignment=TA_CENTER))])
        t = Table(rows, colWidths=[W/6 - 2*mm])
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,-1), bg_color),
            ('TOPPADDING',    (0,0),(-1,-1), 7),
            ('BOTTOMPADDING', (0,0),(-1,-1), 7),
            ('LEFTPADDING',   (0,0),(-1,-1), 3),
            ('RIGHTPADDING',  (0,0),(-1,-1), 3),
            ('BOX',           (0,0),(-1,-1), 0.5, colors.HexColor("#E5E7EB")),
        ]))
        return t

    _has_score = avg_score is not None and not pd.isna(avg_score)
    site_score_pct = f"{avg_score:.0%}" if _has_score else "—"
    site_score_clr = (C_GREEN if avg_score==1.0 else (C_YELLOW if avg_score>=0.7 else C_RED)) if _has_score else C_GREY
    site_score_bg  = (C_GREEN_BG if avg_score==1.0 else (C_YELLOW_BG if avg_score>=0.7 else C_RED_BG)) if _has_score else GP_LIGHT
    audited_sub    = f"of {n_audited}/{n_total} audited" if n_audited < n_total else f"{n_audited} radars audited"

    cards = Table([[
        make_card("SITE SCORE",         site_score_pct, site_score_clr, site_score_bg),
        make_card("🟢 GOOD",           n_good,         C_GREEN,        C_GREEN_BG),
        make_card("⚠️ NEEDS ATTENTION", n_warn,         C_YELLOW,       C_YELLOW_BG),
        make_card("🔴 CRITICAL",        n_crit,         C_RED,          C_RED_BG),
        make_card("TOTAL RADARS",       n_total,        GP_DARK,        GP_LIGHT),
    ]], colWidths=[W/5]*5, hAlign="LEFT")
    cards.setStyle(TableStyle([
        ('VALIGN',        (0,0),(-1,-1), 'TOP'),
        ('LEFTPADDING',   (0,0),(-1,-1), 2),
        ('RIGHTPADDING',  (0,0),(-1,-1), 2),
    ]))
    story.append(cards)
    story.append(Spacer(1, 8*mm))

    # ── Per-radar detail sections ─────────────────────────────────────────────
    story.append(Paragraph("Radar Detail", ps("rsec", fontName="Helvetica-Bold", fontSize=13, textColor=GP_DARK, spaceAfter=6)))

    if site_radars.empty:
        story.append(Paragraph("No radars found for this site.", s_body))
    else:
        for _, row in site_radars.sort_values("_score", na_position='last').iterrows():
            sc      = row["_score"]
            rname   = str(row.get("Radar Name",""))
            rtype   = str(row.get("Type",""))
            la      = row.get("Last Audit","")
            la_str  = pd.to_datetime(la).strftime("%d %b %Y") if pd.notna(la) and str(la) not in ["","nan","NaT"] else "Not audited"

            # Not audited at all — pd.isna catches both None and float('nan')
            if pd.isna(sc):
                not_aud_hdr = Table([[
                    Paragraph(f"<b>{rname}</b>",
                        ps("rn_na", fontName="Helvetica-Bold", fontSize=12, textColor=GP_DARK)),
                    Paragraph(f"{rtype}  ·  Not yet audited",
                        ps("ri_na", fontName="Helvetica", fontSize=9, textColor=C_GREY, alignment=TA_RIGHT)),
                ]], colWidths=[W*0.35, W*0.65])
                not_aud_hdr.setStyle(TableStyle([
                    ('BACKGROUND',    (0,0),(-1,-1), GP_LIGHT),
                    ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
                    ('LEFTPADDING',   (0,0),(-1,-1), 8),
                    ('RIGHTPADDING',  (0,0),(-1,-1), 8),
                    ('TOPPADDING',    (0,0),(-1,-1), 7),
                    ('BOTTOMPADDING', (0,0),(-1,-1), 7),
                    ('BOX',           (0,0),(-1,-1), 0.5, colors.HexColor("#D1D5DB")),
                ]))
                story.append(KeepTogether([not_aud_hdr, Spacer(1, 3*mm)]))
                continue

            status_txt, status_clr, status_bg, _ = score_status(sc)
            sc_str = f"{sc:.0%}"

            # Collect issues — cols are already short names
            crits, warns = [], []
            for c in dqp_cols:
                v = str(row.get(c,""))
                val_clean = v.replace("🔴","").replace("⚠️","").replace("🟢","").strip()
                if "🔴" in v: crits.append((c, val_clean))
                elif "⚠️" in v: warns.append((c, val_clean))

            # Radar header bar — name left, score badge + status right
            hdr_bg   = C_RED if sc < 0.7 else (C_YELLOW if sc < 1.0 else C_GREEN)
            txt_clr  = WHITE
            sub_clr  = colors.HexColor("#FFCDD2" if sc<0.7 else "#FFF8E1" if sc<1.0 else "#C8E6C9")
            radar_hdr_data = [[
                Paragraph(f"<b>{rname}</b>",
                    ps("rn", fontName="Helvetica-Bold", fontSize=12, textColor=WHITE)),
                Paragraph(f"{rtype}  ·  {la_str}",
                    ps("ri2", fontName="Helvetica", fontSize=8, textColor=sub_clr, alignment=TA_CENTER)),
                Paragraph(f"<b>{sc_str}</b>  {status_txt}",
                    ps("rs", fontName="Helvetica-Bold", fontSize=11, textColor=WHITE, alignment=TA_RIGHT)),
            ]]
            radar_hdr_tbl = Table(radar_hdr_data, colWidths=[W*0.35, W*0.35, W*0.30])
            radar_hdr_tbl.setStyle(TableStyle([
                ('BACKGROUND',    (0,0),(-1,-1), hdr_bg),
                ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
                ('LEFTPADDING',   (0,0),(-1,-1), 8),
                ('RIGHTPADDING',  (0,0),(-1,-1), 8),
                ('TOPPADDING',    (0,0),(-1,-1), 7),
                ('BOTTOMPADDING', (0,0),(-1,-1), 7),
            ]))

            # Issues body — crits first, then warnings as observations
            body_rows = []
            if not crits and not warns:
                body_rows.append([
                    Paragraph("✅  All checks passed — no issues found.", s_issue_ok),
                    Paragraph(""),
                ])
            else:
                for name, val in crits:
                    body_rows.append([
                        Paragraph(f"🔴 <b>{name}</b>", ps("ib", fontName="Helvetica-Bold", fontSize=8, textColor=C_RED)),
                        Paragraph(val, s_small),
                    ])
                for name, val in warns:
                    body_rows.append([
                        Paragraph(f"⚠️ {name}", ps("iw", fontName="Helvetica", fontSize=8, textColor=colors.HexColor("#92400E"))),
                        Paragraph(val, s_small),
                    ])

            body_tbl = Table(body_rows, colWidths=[W*0.3, W*0.7])
            body_tbl.setStyle(TableStyle([
                ('BACKGROUND',    (0,0),(-1,-1), status_bg),
                ('VALIGN',        (0,0),(-1,-1), 'TOP'),
                ('LEFTPADDING',   (0,0),(-1,-1), 8),
                ('RIGHTPADDING',  (0,0),(-1,-1), 8),
                ('TOPPADDING',    (0,0),(-1,-1), 5),
                ('BOTTOMPADDING', (0,0),(-1,-1), 5),
                ('LINEBELOW',     (0,0),(-1,-2), 0.3, colors.HexColor("#E5E7EB")),
            ]))

            story.append(KeepTogether([radar_hdr_tbl, body_tbl, Spacer(1, 4*mm)]))

    # ── Recommendations section ───────────────────────────────────────────────
    crits_all = site_radars[site_radars["_score"].notna() & (site_radars["_score"] < 0.7)]  if not site_radars.empty else pd.DataFrame()
    warns_all = site_radars[site_radars["_score"].notna() & (site_radars["_score"] >= 0.7) & (site_radars["_score"] < 1.0)] if not site_radars.empty else pd.DataFrame()

    # Build action lookup from FIXES (short name → action)
    def get_action(col_name):
        for k, v in FIXES.items():
            if k.lower() in col_name.lower() or col_name.lower() in k.lower():
                return v
        return "Contact GroundProbe Geotechnical Support"

    if not crits_all.empty or not warns_all.empty:
        story.append(HRFlowable(width=W, thickness=0.5, color=GP_MID, spaceBefore=4, spaceAfter=6))
        story.append(Paragraph("Recommended Actions",
            ps("rec", fontName="Helvetica-Bold", fontSize=13, textColor=GP_DARK, spaceAfter=3)))
        story.append(Paragraph(
            "The following actions are recommended to improve data quality at this site.",
            ps("rec_sub", fontName="Helvetica", fontSize=9, textColor=C_GREY, spaceAfter=6)))

        # Header row
        rec_rows = [[
            Paragraph("<b>#</b>",       ps("rh0", fontName="Helvetica-Bold", fontSize=8, textColor=WHITE, alignment=TA_CENTER)),
            Paragraph("<b>RADAR</b>",   ps("rh1", fontName="Helvetica-Bold", fontSize=8, textColor=WHITE)),
            Paragraph("<b>FINDING</b>", ps("rh2", fontName="Helvetica-Bold", fontSize=8, textColor=WHITE)),
            Paragraph("<b>RECOMMENDED ACTION</b>", ps("rh3", fontName="Helvetica-Bold", fontSize=8, textColor=WHITE)),
        ]]

        priority = 1
        n_crit_rows = 0

        for _, row in crits_all.iterrows():
            for c in dqp_cols:
                v = str(row.get(c,""))
                if "🔴" in v:
                    finding = v.replace("🔴","").strip()
                    action  = get_action(c)
                    rec_rows.append([
                        Paragraph(f"<b>{priority}</b>",
                            ps(f"pn{priority}", fontName="Helvetica-Bold", fontSize=9, textColor=WHITE, alignment=TA_CENTER)),
                        Paragraph(f"<b>{row['Radar Name']}</b>",
                            ps(f"rr{priority}", fontName="Helvetica-Bold", fontSize=8, textColor=GP_DARK)),
                        Paragraph(f"🔴 {c}: {finding}",
                            ps(f"rf{priority}", fontName="Helvetica", fontSize=8, textColor=C_RED, leading=11)),
                        Paragraph(action,
                            ps(f"ra{priority}", fontName="Helvetica", fontSize=8, textColor=GP_DARK, leading=11)),
                    ])
                    priority += 1
                    n_crit_rows += 1
                    if priority > 20: break
            if priority > 20: break

        for _, row in warns_all.iterrows():
            for c in dqp_cols:
                v = str(row.get(c,""))
                if "🔴" in v or "⚠️" in v:
                    icon    = "🔴" if "🔴" in v else "⚠️"
                    finding = v.replace("🔴","").replace("⚠️","").strip()
                    action  = get_action(c)
                    txt_clr = C_RED if "🔴" in v else colors.HexColor("#92400E")
                    rec_rows.append([
                        Paragraph(f"<b>{priority}</b>",
                            ps(f"pnw{priority}", fontName="Helvetica-Bold", fontSize=9, textColor=WHITE, alignment=TA_CENTER)),
                        Paragraph(str(row['Radar Name']),
                            ps(f"rrw{priority}", fontName="Helvetica", fontSize=8, textColor=GP_DARK)),
                        Paragraph(f"{icon} {c}: {finding}",
                            ps(f"rfw{priority}", fontName="Helvetica", fontSize=8, textColor=txt_clr, leading=11)),
                        Paragraph(action,
                            ps(f"raw{priority}", fontName="Helvetica", fontSize=8, textColor=GP_DARK, leading=11)),
                    ])
                    priority += 1
                    if priority > 30: break
            if priority > 30: break

        if len(rec_rows) > 1:  # more than just header
            rec_tbl = Table(rec_rows, colWidths=[7*mm, W*0.15, W*0.38, W*0.38])
            n_data = len(rec_rows) - 1  # exclude header

            row_styles = [
                # Header
                ('BACKGROUND',    (0,0), (-1,0), GP_DARK),
                ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
                ('LEFTPADDING',   (0,0),(-1,-1), 5),
                ('RIGHTPADDING',  (0,0),(-1,-1), 5),
                ('TOPPADDING',    (0,0),(-1,-1), 5),
                ('BOTTOMPADDING', (0,0),(-1,-1), 5),
                ('LINEBELOW',     (0,1),(-1,-2), 0.3, colors.HexColor("#E5E7EB")),
                ('ROWBACKGROUNDS',(0,1),(-1,-1), [WHITE, GP_LIGHT]),
                # Number badge: red for critical rows, orange for warning
                ('BACKGROUND', (0,1), (0, n_crit_rows), C_RED),
            ]
            if n_crit_rows < n_data:
                row_styles.append(('BACKGROUND', (0, n_crit_rows+1), (0,-1), GP_ORANGE))

            rec_tbl.setStyle(TableStyle(row_styles))
            story.append(rec_tbl)
    # ── Footer ────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 6*mm))
    story.append(HRFlowable(width=W, thickness=0.5, color=GP_MID, spaceAfter=4))

    footer_data = [[
        Paragraph("GroundProbe  ·  groundprobe.com",
            ps("fl", fontSize=7, textColor=C_LGREY)),
        Paragraph(f"Proactive DQP Parameters  ·  {site}  ·  {today_short}",
            ps("fc", fontSize=7, textColor=C_LGREY, alignment=TA_CENTER)),
        Paragraph("Confidential",
            ps("fr", fontSize=7, textColor=C_LGREY, alignment=TA_RIGHT)),
    ]]
    footer_tbl = Table(footer_data, colWidths=[W*0.33, W*0.34, W*0.33])
    footer_tbl.setStyle(TableStyle([
        ('VALIGN',  (0,0),(-1,-1), 'MIDDLE'),
        ('LEFTPADDING',  (0,0),(-1,-1), 0),
        ('RIGHTPADDING', (0,0),(-1,-1), 0),
    ]))
    story.append(footer_tbl)

    doc.build(story)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────────────────────

def _render_sidebar(page: str):
    """Render sidebar header + fleet stats for all non-Audit pages."""
    with st.sidebar:
        st.markdown(
            '<div style="padding:4px 0 10px 0;border-bottom:1px solid #2D4068;margin-bottom:10px">'
            '<div style="font-size:1.1rem;font-weight:700;color:#F78F1E">🛰️ Fleet DQP</div>'
            '<div style="font-size:.7rem;color:#7A9BBF;margin-top:2px;letter-spacing:.03em">'
            'GROUNDPROBE · DATA QUALITY PARAMETERS</div></div>',
            unsafe_allow_html=True
        )
        try:
            _wb     = get_workbook()
            _mat    = get_audit_matrix(_wb)
            _hist   = get_history(_wb)
            _scored = _hist[_hist["Score Captured"].notna()] if not _hist.empty else pd.DataFrame()
            _n      = len(_mat)
            _aud    = int(_scored["Radar Name"].nunique()) if not _scored.empty else 0
            _cov    = f"{_aud/_n:.0%}" if _n else "0%"
            _avg    = _scored["Score Captured"].mean() if not _scored.empty else None
            _avg_s  = f"{_avg:.0%}" if _avg is not None and pd.notna(_avg) else "—"
            _bus    = int(_mat["BU"].nunique()) if "BU" in _mat.columns else 0

            _grid = (
                '<div style="font-size:.62rem;color:#7A9BBF;text-transform:uppercase;'
                'letter-spacing:.06em;margin-bottom:5px">Fleet Status</div>'
                '<div style="display:grid;grid-template-columns:1fr 1fr;gap:5px;margin-bottom:10px">'
            )
            for _l, _v in [("Radars", _n), ("BUs", _bus), ("Coverage", _cov), ("Avg Score", _avg_s)]:
                _grid += (
                    f'<div style="background:#152236;border-radius:5px;padding:7px 9px">'
                    f'<div style="font-size:.6rem;color:#7A9BBF">{_l}</div>'
                    f'<div style="font-size:1rem;font-weight:700;color:#F78F1E">{_v}</div>'
                    f'</div>'
                )
            _grid += '</div>'
            st.markdown(_grid, unsafe_allow_html=True)

            # Page-specific extras — only what the page itself doesn't already show
            if "Expir" in page:
                _now = pd.Timestamp.now()
                _exp = 0
                if not _scored.empty:
                    _exp = int(_scored.groupby("Radar Name")["Audit Date"].max().apply(
                        lambda d: (_now - pd.to_datetime(d)).days > 180 if pd.notna(d) else False).sum())
                st.markdown(
                    f'<div style="font-size:.8rem;line-height:1.9">'
                    f'🔴 <b style="color:#FF8080">{_exp}</b> expired<br>'
                    f'⬛ <b style="color:#CDD5E0">{_n - _aud}</b> never audited<br>'
                    f'🟢 <b style="color:#6BCB77">{_aud - _exp}</b> up to date'
                    f'</div>',
                    unsafe_allow_html=True
                )
            elif "History" in page:
                _tot = len(_hist) if not _hist.empty else 0
                st.caption(f"📋 {_tot} audit records · {_aud} radars audited")
            elif "Fleet Management" in page:
                _no_ra = int((_mat["Remote Access"].isna() | (_mat["Remote Access"].astype(str).str.strip() == "")).sum()) if "Remote Access" in _mat.columns else 0
                if _no_ra > 0:
                    st.warning(f"⚠️ {_no_ra} radars missing Remote Access")
        except Exception:
            pass
        st.divider()

st.set_page_config(page_title="Fleet DQP · GroundProbe", page_icon="🛰️", layout="wide", initial_sidebar_state="expanded")
# Custom favicon injection
st.markdown("""
<head>
<link rel="icon" type="image/png" href="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAEAAAABACAYAAACqaXHeAAAG7UlEQVR4nO2aS2wbxxmA/39md/mUKJKiRMkiRct6MJYjxZKM2ogfstuitwK9GegjbR1bRQtUKNqkCFoUbZK6CNIgNuCmSvNA2jhRmgA9NE6KopfYiG1ZNuRSiiJa1is2SUkwKfFNLndnpgfDhZsY7nEO3u+2cxh882FmZw+Lex49KOABhsgWkI0VQLaAbKwAsgVkYwWQLSAbK4BsAdlYAWQLyMYKIFtANlYA2QKysQLIFpCNFUC2gGysALIFZGMFkC0gGyuAbAHZWAFkC8hGuddgu9ucP9ZXyfc0mA8hcLXeTvDklOPimRX7foWAYXJQEUEgAOcCKEHBBQAIgXcHFQCAdx4ICs4FEkQQQgASFJwimAZH7fEd5bNBm+48ftU3+Ne33kg4XG4HoYSeOfPhzNjLr+67MwfngiiKYmzdGlmqc7sruXzetbi41AUA8M74X1ZOnx5PnPngH3tVVakZhqkRQhjnnCKiYKaZNAw9Y7M7++8boNdnxI7vLXVe3yCVE1dd08ky9Xk1UV6vYBsAgMlBBQAQAlAAUEKAcY707jnuLPLuMX47zn/HuUDCBWgEgXU3sMYbG2yzvq4+uyUUCj/x5C8uO5wO4+lf//LA+fMT07HpmT4AAEVRjOED+64oisJN0ySBpkDJ7/fl4/FrHaG2Le3JZCoHAGAYpnY7GKc2m62i67qDUNqmEXv9/90BTw4Ug/EMXfz5+fp+AAC7Isqf5VHjApSjvaVzPV7WoFHmfG3WXdjfqhcHWnjoVhGzL0y5/HmDeH+/N5c6fsWtuhRR/mFfUfnZxw0dP+4vXm50CIdGuFaoYeU3k/VDIbe5NLqzXPBo3LfVS3o+vAYToXBbUgjhTSST/mhP9yoAgN1hN154/neXXW6XXXAhxl5+1eX1NpQNw6SMMWzwePT9+/bO6npt4AcjjxOPp375xZOnNi5dujw4cuzIR12d21wAAC+ePNWcTKbCn1/vF94BXY2k+aOElgUA+Ma2ytk3v1bIvnJo85ZbFdnBJqNdAIh3rjnTh7srpK/RiDx70Vl1qtz9/d5Syqux9I6A6M7qxN/jNdIBB29mHJRHW/Sh5RzJv7fgLO4P8aGAg6ee2lUgOkPzxFXXhs4EXSuonnA4lAMAeObpX5lHj34vfOoPY2dv3kg07dnzpV2x2EzmzbfG9dj0TF9seqZz4tLkTs4Zmf10rk1RFKFpqu2lsT+JRCKZHjl2xHno4IGLj33nm8Pz1xdKQ0MDD3/3sW/dgNvH8v4B5tI8+fVtekuDja+/v2zf/a/P6IJNEQ6Tg+q3saZ35+36hTXb7pDbDE+saUuLeSVaMaBKEDDiYWvpEquVDKyPek1tIUuXnKooOFXU/rbo6NEZKGsFliMIvLcRI6fjdiW+qfaqBDFVcQQfikb5aiKRGBn50ZbDh78dfHv83QMdW8Orhq7rfxx7ZfjChYld3d1dcc45IQT5wuJSKJlMhYPBJraysrISi830mczg5WJR790erWUymfRcPO549rfPTf39/Q88AP97LO95BJ6ZrCv9dLDMXjqYJ+slXO70iYGPE+onNgrdNhW9q2XaCADw+pxz9sj2ajTqy8eCLhF47pzTF7CLjKYgGX2kcH44xHa+MatNem0szQW4M1US/HKoGs/VyOZ6mUTenlPPPTVU7shUYf5WSbTkUQt0dnaszS8uJytVPXTHJxRur+QKxSwhpPGrXzk0wThHv99XoISISrWqZrM5X3Nzs9rYGPCdPPH8VDgUCo3+5AndNE3/8PCB6q6hIb2uzk1fe/3PXzj/AAB4rz9EWl1suT9gJrw2zmoccSlLXSUDHREP27iepQHGUalxULs8ZrLdI2pT69S7qROPSsHc7jNSXCDWGJDNKnHUOChtbpaLb6otrS6W0YhgKwWliQmg231mIuhiLFGg6mJOaY5Gu9fy+YK2urrmJQiCcUHCba2ZGnXRyU+WhsKhtpWH+3YkioWiSigVdXVu48qVqUikAVPbIuEKEAL/vhrzpjMZDyKK1taWzODO/tLN1TT558SnezjnVAiB9w3gUkVu9JFi3EaBVk1gBASxKaBQBFJlWHMqQgMhEBBQZ1irMjRdKtdUBIUDioqJOgAAQYEKCooAqHNiOCm31TiaAkDYiFARAUsm6gZHZqOC2onQKpVKjVKFaJqqiNtXiXAQxt67RmsX057dAAB+v389EgmnCKK4cTMRzGXWPaO9mRkNdFIzhXC5HDZN1SgAQFXXa8TUzbmsVhlfCQwjCv65q/reO+BB4oH/ErQCyBaQjRVAtoBsrACyBWRjBZAtIBsrgGwB2VgBZAvIxgogW0A2VgDZArKxAsgWkI0VQLaAbKwAsgVkYwWQLSAbK4BsAdn8B9HwG7oXs3PoAAAAAElFTkSuQmCC">
<link rel="shortcut icon" type="image/png" href="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAEAAAABACAYAAACqaXHeAAAG7UlEQVR4nO2aS2wbxxmA/39md/mUKJKiRMkiRct6MJYjxZKM2ogfstuitwK9GegjbR1bRQtUKNqkCFoUbZK6CNIgNuCmSvNA2jhRmgA9NE6KopfYiG1ZNuRSiiJa1is2SUkwKfFNLndnpgfDhZsY7nEO3u+2cxh882FmZw+Lex49KOABhsgWkI0VQLaAbKwAsgVkYwWQLSAbK4BsAdlYAWQLyMYKIFtANlYA2QKysQLIFpCNFUC2gGysALIFZGMFkC0gGyuAbAHZWAFkC8hGuddgu9ucP9ZXyfc0mA8hcLXeTvDklOPimRX7foWAYXJQEUEgAOcCKEHBBQAIgXcHFQCAdx4ICs4FEkQQQgASFJwimAZH7fEd5bNBm+48ftU3+Ne33kg4XG4HoYSeOfPhzNjLr+67MwfngiiKYmzdGlmqc7sruXzetbi41AUA8M74X1ZOnx5PnPngH3tVVakZhqkRQhjnnCKiYKaZNAw9Y7M7++8boNdnxI7vLXVe3yCVE1dd08ky9Xk1UV6vYBsAgMlBBQAQAlAAUEKAcY707jnuLPLuMX47zn/HuUDCBWgEgXU3sMYbG2yzvq4+uyUUCj/x5C8uO5wO4+lf//LA+fMT07HpmT4AAEVRjOED+64oisJN0ySBpkDJ7/fl4/FrHaG2Le3JZCoHAGAYpnY7GKc2m62i67qDUNqmEXv9/90BTw4Ug/EMXfz5+fp+AAC7Isqf5VHjApSjvaVzPV7WoFHmfG3WXdjfqhcHWnjoVhGzL0y5/HmDeH+/N5c6fsWtuhRR/mFfUfnZxw0dP+4vXm50CIdGuFaoYeU3k/VDIbe5NLqzXPBo3LfVS3o+vAYToXBbUgjhTSST/mhP9yoAgN1hN154/neXXW6XXXAhxl5+1eX1NpQNw6SMMWzwePT9+/bO6npt4AcjjxOPp375xZOnNi5dujw4cuzIR12d21wAAC+ePNWcTKbCn1/vF94BXY2k+aOElgUA+Ma2ytk3v1bIvnJo85ZbFdnBJqNdAIh3rjnTh7srpK/RiDx70Vl1qtz9/d5Syqux9I6A6M7qxN/jNdIBB29mHJRHW/Sh5RzJv7fgLO4P8aGAg6ee2lUgOkPzxFXXhs4EXSuonnA4lAMAeObpX5lHj34vfOoPY2dv3kg07dnzpV2x2EzmzbfG9dj0TF9seqZz4tLkTs4Zmf10rk1RFKFpqu2lsT+JRCKZHjl2xHno4IGLj33nm8Pz1xdKQ0MDD3/3sW/dgNvH8v4B5tI8+fVtekuDja+/v2zf/a/P6IJNEQ6Tg+q3saZ35+36hTXb7pDbDE+saUuLeSVaMaBKEDDiYWvpEquVDKyPek1tIUuXnKooOFXU/rbo6NEZKGsFliMIvLcRI6fjdiW+qfaqBDFVcQQfikb5aiKRGBn50ZbDh78dfHv83QMdW8Orhq7rfxx7ZfjChYld3d1dcc45IQT5wuJSKJlMhYPBJraysrISi830mczg5WJR790erWUymfRcPO549rfPTf39/Q88AP97LO95BJ6ZrCv9dLDMXjqYJ+slXO70iYGPE+onNgrdNhW9q2XaCADw+pxz9sj2ajTqy8eCLhF47pzTF7CLjKYgGX2kcH44xHa+MatNem0szQW4M1US/HKoGs/VyOZ6mUTenlPPPTVU7shUYf5WSbTkUQt0dnaszS8uJytVPXTHJxRur+QKxSwhpPGrXzk0wThHv99XoISISrWqZrM5X3Nzs9rYGPCdPPH8VDgUCo3+5AndNE3/8PCB6q6hIb2uzk1fe/3PXzj/AAB4rz9EWl1suT9gJrw2zmoccSlLXSUDHREP27iepQHGUalxULs8ZrLdI2pT69S7qROPSsHc7jNSXCDWGJDNKnHUOChtbpaLb6otrS6W0YhgKwWliQmg231mIuhiLFGg6mJOaY5Gu9fy+YK2urrmJQiCcUHCba2ZGnXRyU+WhsKhtpWH+3YkioWiSigVdXVu48qVqUikAVPbIuEKEAL/vhrzpjMZDyKK1taWzODO/tLN1TT558SnezjnVAiB9w3gUkVu9JFi3EaBVk1gBASxKaBQBFJlWHMqQgMhEBBQZ1irMjRdKtdUBIUDioqJOgAAQYEKCooAqHNiOCm31TiaAkDYiFARAUsm6gZHZqOC2onQKpVKjVKFaJqqiNtXiXAQxt67RmsX057dAAB+v389EgmnCKK4cTMRzGXWPaO9mRkNdFIzhXC5HDZN1SgAQFXXa8TUzbmsVhlfCQwjCv65q/reO+BB4oH/ErQCyBaQjRVAtoBsrACyBWRjBZAtIBsrgGwB2VgBZAvIxgogW0A2VgDZArKxAsgWkI0VQLaAbKwAsgVkYwWQLSAbK4BsAdn8B9HwG7oXs3PoAAAAAElFTkSuQmCC">
</head>
""", unsafe_allow_html=True)
st.markdown("""
<style>
  /* ── GroundProbe Brand Theme ── */
  /* Navy #1B2A4A | Orange #F78F1E | Dark navy #112038 | Light bg #F4F6F9 */

  .block-container{padding-top:1rem;max-width:1400px}

  /* Sidebar — GroundProbe navy */
  [data-testid="stSidebar"]{background:#1B2A4A!important}
  [data-testid="stSidebar"] p,
  [data-testid="stSidebar"] span,
  [data-testid="stSidebar"] label,
  [data-testid="stSidebar"] div{color:#E8EDF5!important}
  [data-testid="stSidebar"] h1,
  [data-testid="stSidebar"] h2,
  [data-testid="stSidebar"] h3{color:#F78F1E!important;font-weight:700!important}
  [data-testid="stSidebar"] hr{border-color:#2D4068!important;opacity:1!important}
  [data-testid="stSidebar"] [data-baseweb="select"] *{background:#243D63!important;color:#E8EDF5!important;border-color:#3D5585!important}
  [data-testid="stSidebar"] [data-baseweb="select"] svg{color:#A8BBCF!important;fill:#A8BBCF!important}
  [data-testid="stSidebar"] input{background:#243D63!important;border-color:#3D5585!important;color:#E8EDF5!important}
  [data-testid="stSidebar"] input::placeholder{color:#7A9BBF!important}
  [data-testid="stSidebar"] [data-testid="stExpander"]{background:#152236!important;border:1px solid #2D4068!important}
  [data-testid="stSidebar"] [data-testid="stExpander"] summary p{color:#E8EDF5!important}
  [data-testid="stSidebar"] .stMarkdown code{background:#243D63!important;color:#F78F1E!important;border:none!important}
  [data-testid="stSidebar"] strong{color:#FFFFFF!important}
  [data-testid="stSidebar"] small, [data-testid="stSidebar"] .stCaption{color:#7A9BBF!important}
  [data-testid="stSidebar"] [data-testid="stMetricValue"]{color:#FFFFFF!important}
  [data-testid="stSidebar"] [role="option"]{background:#1B2A4A!important;color:#E8EDF5!important}
  [data-testid="stSidebar"] [role="option"]:hover{background:#2D4068!important}
  [data-testid="stSidebar"] [data-testid="stCheckbox"] span{color:#E8EDF5!important}
  [data-testid="stSidebar"] button{background:#243D63!important;border-color:#3D5585!important;color:#E8EDF5!important}
  [data-testid="stSidebar"] button:hover{background:#2D4068!important}
  /* Status indicators stay colored */
  [data-testid="stSidebar"] .stAlert{background:#243D63!important}

  /* Main area */
  .main .block-container{background:#F4F6F9}

  /* Page headers — orange accent */
  h1,h2{color:#1B2A4A!important;font-weight:700!important;border-left:4px solid #F78F1E;padding-left:12px}
  h3,h4{color:#1B2A4A!important;font-weight:600!important}

  /* Radio buttons */
  .stRadio>div{flex-direction:row;gap:.4rem;flex-wrap:wrap}
  .stRadio label{font-size:.82rem!important}
  div[data-testid="stRadio"]>label{font-weight:600;font-size:.9rem;color:#1B2A4A}

  /* Tab active indicator — orange */
  .stTabs [data-baseweb="tab-highlight"]{background:#F78F1E!important}
  .stTabs [data-baseweb="tab"]{font-weight:500;color:#4A5568}
  .stTabs [aria-selected="true"]{color:#1B2A4A!important;font-weight:700!important}

  /* Primary buttons — GroundProbe orange */
  .stButton button[kind="primary"]{background:#F78F1E!important;border-color:#F78F1E!important;color:white!important;font-weight:600!important}
  .stButton button[kind="primary"]:hover{background:#E07A0A!important;border-color:#E07A0A!important}

  /* Metric cards */
  [data-testid="stMetric"]{background:white;border-radius:8px;padding:12px 16px;border:1px solid #E2E8F0;box-shadow:0 1px 3px rgba(27,42,74,.06)}
  [data-testid="stMetricLabel"]{color:#64748B!important;font-size:.75rem!important;font-weight:600!important;text-transform:uppercase!important;letter-spacing:.04em!important}
  [data-testid="stMetricValue"]{color:#1B2A4A!important;font-weight:800!important}

  /* Dataframes */
  [data-testid="stDataFrame"] thead tr th{background:#1B2A4A!important;color:white!important;font-weight:600!important;font-size:.78rem!important;text-transform:uppercase!important;letter-spacing:.04em!important}

  /* Dividers */
  hr{border-color:#E2E8F0}

  /* Expanders */
  [data-testid="stExpander"]{border:1px solid #E2E8F0!important;border-radius:8px!important;background:white!important}
  [data-testid="stExpander"] summary{font-weight:600;color:#1B2A4A}

  /* Info/warning/success boxes */
  [data-testid="stAlert"]{border-radius:8px!important}

  /* Custom components */
  .cat-header{background:#1B2A4A;color:white;padding:5px 14px;border-radius:6px;font-weight:700;font-size:.95rem;margin:1rem 0 .4rem 0}
  .score-green{background:#00875A;color:white;padding:12px 20px;border-radius:8px;font-size:2rem;font-weight:700;text-align:center}
  .score-yellow{background:#F78F1E;color:white;padding:12px 20px;border-radius:8px;font-size:2rem;font-weight:700;text-align:center}
  .score-red{background:#D32F2F;color:white;padding:12px 20px;border-radius:8px;font-size:2rem;font-weight:700;text-align:center}
  .na-field{opacity:.5;padding:3px 10px;border-radius:4px;font-size:.8rem;display:inline-block;margin:2px 0;border:1px solid currentColor}
  .albert-banner{border-left:4px solid #F78F1E;background:#FFF8F0;padding:9px 14px;border-radius:4px;margin:6px 0;font-size:.84rem}
  .hist-entry{font-size:.8rem;opacity:.75;padding:1px 0}

  /* KPI cards with brand colors */
  .dqp-kpi{border-left:4px solid var(--kpi-color,#F78F1E);border-radius:8px;padding:14px 18px;margin:2px 0;background:white;box-shadow:0 1px 3px rgba(27,42,74,.06)}
  .dqp-kpi .kpi-label{font-size:.72rem;font-weight:700;text-transform:uppercase;letter-spacing:.05em;color:var(--kpi-color,#F78F1E)}
  .dqp-kpi .kpi-value{font-size:2rem;font-weight:800;line-height:1.1;color:#1B2A4A}
  .dqp-kpi .kpi-sub{font-size:.78rem;color:#64748B;margin-top:2px}

  /* Page title accent bar */
  .gp-page-header{border-left:4px solid #F78F1E;padding-left:12px;margin-bottom:4px}
</style>
""", unsafe_allow_html=True)

page = st.sidebar.selectbox("", ["🏠 Executive Dashboard", "🛰️ Audit Radar", "📊 History & Trends", "⏰ Expirations", "📄 Client PDF Report", "⚙️ Fleet Management"], label_visibility="collapsed")


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE 0 — EXECUTIVE DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
if "Executive" in page or "Dashboard" in page:
    _render_sidebar(page)
    wb     = get_workbook()
    df_mat = get_audit_matrix(wb)
    df_hist = get_history(wb)

    st.markdown('## 🏠 Fleet Executive Dashboard')
    st.caption(f"GroundProbe · {len(df_mat)} radars · {df_mat['Site'].nunique()} sites · {df_mat['BU'].nunique() if 'BU' in df_mat.columns else '—'} Business Units")
    st.divider()

    # ── Build last audit — HISTORY_LOG only, scored entries only ──────────────────
    if not df_hist.empty:
        _hist_sc = df_hist[df_hist["Score Captured"].notna()].dropna(subset=["Audit Date"])
        if not _hist_sc.empty:
            _last = (_hist_sc.sort_values("Audit Date", ascending=False)
                     .groupby("Radar Name").first().reset_index()
                     [["Radar Name","Audit Date","Score Captured","Auditor"]])
        else:
            _last = pd.DataFrame(columns=["Radar Name","Audit Date","Score Captured","Auditor"])
    else:
        _last = pd.DataFrame(columns=["Radar Name","Audit Date","Score Captured","Auditor"])

    _today  = pd.Timestamp.now()
    _merged = df_mat.copy()
    _merged = _merged.merge(_last, on="Radar Name", how="left")
    _merged["Days Since"] = _merged["Audit Date"].apply(
        lambda d: (_today - pd.to_datetime(d)).days if pd.notna(d) else None)

    def _status(row):
        ds = row["Days Since"]
        sc = row.get("Score Captured")
        if ds is None:             return "⬛ Never Audited"
        if ds > 180:               return "🔴 Expired"
        if ds >= 150:              return "⚠️ Expiring Soon"
        if pd.isna(sc):            return "🟢 Audited"
        if sc == 1.0:              return "🟢 Good"
        if sc >= 0.7:              return "⚠️ Needs Attention"
        return "🔴 Critical"

    _merged["Status"] = _merged.apply(_status, axis=1)

    n_total    = len(_merged)
    n_audited  = _merged["Audit Date"].notna().sum()
    n_expired  = (_merged["Days Since"] > 180).sum()
    n_never    = _merged["Audit Date"].isna().sum()
    n_expiring = ((_merged["Days Since"] >= 150) & (_merged["Days Since"] <= 180)).sum()
    coverage   = n_audited / n_total if n_total else 0
    avg_sc     = _last["Score Captured"].mean() if not _last.empty else None

    # ── Global KPIs ───────────────────────────────────────────────────────────
    g1,g2,g3,g4,g5,g6 = st.columns(6)
    g1.metric("Total Radars",    n_total)
    g2.metric("Fleet Coverage",  f"{coverage:.0%}", help="% radars with at least one audit")
    g3.metric("Avg Score",       f"{avg_sc:.0%}" if avg_sc and pd.notna(avg_sc) else "—")
    g4.metric("🔴 Expired",      n_expired,  help=">180 days since last audit")
    g5.metric("⚠️ Expiring Soon", n_expiring, help="150–180 days")
    g6.metric("⬛ Never Audited", n_never)

    st.divider()

    # ── BU Breakdown ──────────────────────────────────────────────────────────
    if "BU" in _merged.columns:
        st.markdown("#### Business Unit Overview")
        bu_rows = []
        for _bu in sorted(_merged["BU"].dropna().unique()):
            _sub       = _merged[_merged["BU"] == _bu]
            _n         = len(_sub)
            _aud       = _sub["Audit Date"].notna().sum()
            _exp       = (_sub["Days Since"] > 180).sum()
            _nev       = _sub["Audit Date"].isna().sum()
            _cov       = _aud / _n if _n else 0
            _scores    = _sub["Score Captured"].dropna()
            _avg_bu    = _scores.mean() if len(_scores) else None
            _crit      = (_scores < 0.7).sum() if len(_scores) else 0
            _warn      = ((_scores >= 0.7) & (_scores < 1.0)).sum() if len(_scores) else 0
            _good      = (_scores == 1.0).sum() if len(_scores) else 0
            bu_rows.append({
                "BU":            _bu,
                "Radars":        _n,
                "Audited":       _aud,
                "Coverage":      f"{_cov:.0%}",
                "Avg Score":     f"{_avg_bu:.0%}" if _avg_bu and pd.notna(_avg_bu) else "—",
                "🟢 Good":        _good,
                "⚠️ Attention":   _warn,
                "🔴 Critical":    _crit,
                "🔴 Expired":     _exp,
                "⚫ Never":       _nev,
            })
        _bu_df = pd.DataFrame(bu_rows)
        st.dataframe(_bu_df, use_container_width=True, hide_index=True)

        st.divider()
        st.markdown("#### Coverage & Score by BU")
        _bc1, _bc2 = st.columns(2)

        with _bc1:
            st.markdown("**Coverage %**")
            _cov_data = {r["BU"]: int(r["Coverage"].strip("%")) for r in bu_rows}
            _cov_df   = pd.DataFrame({"BU": list(_cov_data.keys()), "Coverage %": list(_cov_data.values())})
            st.bar_chart(_cov_df.set_index("BU"), use_container_width=True, height=250)

        with _bc2:
            st.markdown("**Avg Score %**")
            _sc_data = {r["BU"]: int(r["Avg Score"].strip("%")) if r["Avg Score"] != "—" else 0 for r in bu_rows}
            _sc_df   = pd.DataFrame({"BU": list(_sc_data.keys()), "Avg Score %": list(_sc_data.values())})
            st.bar_chart(_sc_df.set_index("BU"), use_container_width=True, height=250)

    st.divider()

    # ── Radar Health Heatmap by BU ────────────────────────────────────────────
    st.markdown("#### Radar Status Summary")
    _s1,_s2,_s3,_s4 = st.columns(4)
    _s1.metric("🟢 Good / Audited",   (_merged["Status"].isin(["🟢 Good","🟢 Audited"])).sum())
    _s2.metric("⚠️ Needs Attention",  (_merged["Status"] == "⚠️ Needs Attention").sum())
    _s3.metric("🔴 Critical",         (_merged["Status"] == "🔴 Critical").sum())
    _s4.metric("🔴 Expired",          (_merged["Status"] == "🔴 Expired").sum())

    st.divider()

    # ── Top 10 sites needing attention ───────────────────────────────────────
    st.markdown("#### Sites Needing Attention")
    _at_risk = _merged[_merged["Status"].isin(["🔴 Critical","🔴 Expired","⚠️ Needs Attention","⚠️ Expiring Soon"])]
    if _at_risk.empty:
        st.success("✅ All audited radars are in good standing.")
    else:
        _site_risk = (_at_risk.groupby(["Site","BU"] if "BU" in _at_risk.columns else ["Site"])
                     .size().reset_index(name="At-Risk Radars")
                     .sort_values("At-Risk Radars", ascending=False)
                     .head(15))
        st.dataframe(_site_risk, use_container_width=True, hide_index=True)

    st.divider()
    st.caption(f"Last refreshed: {_today.strftime('%d %b %Y %H:%M')} · Data from Fleet_DQP_Master.xlsx")

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE 1 — AUDIT
# ═══════════════════════════════════════════════════════════════════════════════
elif "Audit" in page:
    with st.sidebar:
        st.markdown("""
        <div style="padding:4px 0 12px 0;border-bottom:1px solid #2D4068;margin-bottom:12px">
            <div style="font-size:1.1rem;font-weight:700;color:#F78F1E;letter-spacing:.02em">🛰️ Fleet DQP</div>
            <div style="font-size:.72rem;color:#7A9BBF;margin-top:2px;letter-spacing:.03em">GROUNDPROBE · DATA QUALITY PARAMETERS</div>
        </div>
        """, unsafe_allow_html=True)

        wb        = get_workbook()
        df_matrix = get_audit_matrix(wb)
        radar_list = sorted(df_matrix["Radar Name"].dropna().astype(str).tolist(), key=_natural_key)

        # Build label: "SSR126 · XT", "SSR539 · FX", "SAR407 · SAR-X", "SOM505 · SOM"
        def radar_label(r):
            row_data = df_matrix[df_matrix["Radar Name"]==r]
            t = str(row_data["Type"].values[0]).strip() if len(row_data) > 0 else ""
            s = str(row_data["Site"].values[0]).strip() if len(row_data) > 0 else ""
            parts = [r]
            if t and t != "nan": parts.append(t)
            if s and s != "nan": parts.append(s)
            return " · ".join(parts)
        label_map  = {r: radar_label(r) for r in radar_list}
        label_to_r = {v: k for k, v in label_map.items()}

        search = st.text_input("🔍 Search radar or site", placeholder="SSR329 / Arcelormittal")

        # BU filter
        all_bus = sorted(df_matrix["BU"].dropna().unique().tolist())
        if len(all_bus) > 1:
            bu_filter = st.selectbox("🌎 Business Unit", ["All"] + all_bus, key="bu_filter")
        else:
            bu_filter = "All"

        if search:
            s = search.upper()
            filtered = [r for r in radar_list if s in r.upper() or
                        s in str(df_matrix.loc[df_matrix["Radar Name"]==r,"Site"].values[0]).upper()]
        else:
            filtered = radar_list

        if bu_filter != "All":
            filtered = [r for r in filtered
                        if df_matrix.loc[df_matrix["Radar Name"]==r,"BU"].values[0] == bu_filter]

        if not filtered:
            st.warning("No radars found.")
            st.stop()

        filtered_labels = [label_map[r] for r in filtered]
        selected_label  = st.selectbox("Select Radar", filtered_labels)
        selected        = label_to_r[selected_label]
        radar_row = df_matrix[df_matrix["Radar Name"]==selected].iloc[0]
        site_val  = str(radar_row.get("Site",""))
        type_val  = str(radar_row.get("Type","")).upper()
        pa_stl    = str(radar_row.get("PA","N/A") or "N/A").strip()
        dim_label = "3D · XT" if is_3d(type_val) else "2D · FX/SOM/SAR"

        st.divider()
        bu_val = str(radar_row.get("BU","") or "")
        # Load history early — needed for the Never Audited badge
        df_hist = get_history(wb)
        _rname_check = str(radar_row.get("Radar Name",""))
        _has_audit   = (not df_hist.empty and _rname_check in df_hist["Radar Name"].values)
        if not _has_audit:
            st.markdown(
                '<div style="background:#F78F1E;color:#fff;border-radius:6px;'
                'padding:3px 10px;font-size:.75rem;font-weight:700;letter-spacing:.03em;'
                'display:inline-block;margin-bottom:4px">⬛ NEVER AUDITED</div>',
                unsafe_allow_html=True)
        st.markdown(f"**📍 Site:** {site_val}")
        if bu_val and bu_val != "Other":
            st.markdown(f"**🌎 BU:** `{bu_val}`")
        st.markdown(f"**📡 Type:** `{type_val}` — {dim_label}")

        # Remote Access badge
        remote = str(radar_row.get("Remote Access","") or "").strip()
        if remote == "Customer":
            st.markdown(f"**📞 Access:** 🟡 `Customer` — Teams call required")
        elif remote:
            st.markdown(f"**🌐 Access:** 🟢 `{remote}`")

        # ── Real audit stats from HISTORY_LOG (already loaded above) ───────────
        radar_hist = df_hist[df_hist["Radar Name"]==selected].sort_values("Audit Date", ascending=False)

        real_total = len(radar_hist)
        if not radar_hist.empty and pd.notna(radar_hist.iloc[0]["Audit Date"]):
            real_last = pd.to_datetime(radar_hist.iloc[0]["Audit Date"])
            days_ago  = (datetime.now() - real_last).days
            icon      = "🟢" if days_ago <= 180 else "🔴"
            st.markdown(f"**{icon} Last Audit:** {real_last.strftime('%d %b %Y')} ({days_ago}d ago)")
            if days_ago > 180:
                st.markdown("**⚠️ EXPIRED** — audit overdue (>180 days)")
        else:
            st.markdown("**⚪ Last Audit:** Never audited")

        st.markdown(f"**📊 Total Audits:** {real_total}")

        if not radar_hist.empty:
            st.divider()
            st.markdown("**📋 Audit History**")
            for _, h in radar_hist.head(5).iterrows():
                d  = h["Audit Date"].strftime("%d/%m/%y") if pd.notna(h["Audit Date"]) else "?"
                sc = f"{h['Score Captured']:.0%}" if pd.notna(h["Score Captured"]) else "?"
                au = h.get("Auditor","?") or "?"
                icon = "🟢" if h["Score Captured"]==1 else ("⚠️" if (h["Score Captured"] or 0)>=0.7 else "🔴") if pd.notna(h["Score Captured"]) else "⚪"
                st.markdown(f"<div class='hist-entry'>{icon} {d} · {au} · {sc}</div>", unsafe_allow_html=True)

        st.divider()
        st.caption("🟢 Connected · GroundProbe Advisory Services")


    # ── Main ──────────────────────────────────────────────────────────────────
    st.markdown(f"## 🛰️ `{selected}` — {site_val}")
    st.caption(f"Type: **{type_val}** · {dim_label}")

    # Pre-fill info banner — based on real HISTORY_LOG (load early, used for auditor dropdown too)
    df_hist_pf    = get_history(wb)
    radar_hist_pf = df_hist_pf[df_hist_pf["Radar Name"]==selected].sort_values("Audit Date", ascending=False)

    c1, c2 = st.columns(2)
    # Suggest known auditors from history
    _known_auditors = []
    if not df_hist_pf.empty:
        _known_auditors = sorted(df_hist_pf["Auditor"].dropna().unique().tolist())
    with c1:
        if _known_auditors:
            _aud_opts = [""] + _known_auditors + ["Other (type below)"]
            _aud_sel  = st.selectbox("👤 Auditor", _aud_opts,
                                      format_func=lambda x: "Select auditor..." if x == "" else x)
            if _aud_sel == "Other (type below)" or _aud_sel == "":
                auditor_input = st.text_input("Type auditor name", placeholder="Your name")
            else:
                auditor_input = _aud_sel
        else:
            auditor_input = st.text_input("👤 Auditor", placeholder="Your name")
    with c2:
        audit_date_input = st.date_input("📅 Date", value=date.today())
    notes_input = st.text_area("📝 Notes / Observations", placeholder="Optional — describe any field conditions, access issues, or context for this audit...", height=80)

    # Pre-fill banner
    has_prev = not radar_hist_pf.empty
    if has_prev:
        prev_date  = radar_hist_pf.iloc[0]["Audit Date"]
        prev_score = radar_hist_pf.iloc[0]["Score Captured"]
        days_ago   = (datetime.now() - pd.to_datetime(prev_date)).days if pd.notna(prev_date) else None
        days_str   = f"{days_ago}d ago" if days_ago is not None else "?"
        score_str  = f"{prev_score:.0%}" if pd.notna(prev_score) else "?"
        date_str   = prev_date.strftime('%d %b %Y') if pd.notna(prev_date) else "?"
        st.info(f"📋 **Pre-filled from last audit** ({date_str}, {days_str}, score {score_str}). Change only what's different.")
    else:
        st.info("⚪ **First audit for this radar** — all fields start at default. Fill in what you see.")

    st.divider()

    answers     = {}
    current_cat = None
    for key in DQP_MASTER:
        cat   = key.split("|")[0].strip()
        param = key.split("|")[1].strip() if "|" in key else key
        opts  = get_options(key, type_val, pa_stl)

        if cat != current_cat:
            current_cat = cat
            st.markdown(f"<div class='cat-header'>📂 {cat}</div>", unsafe_allow_html=True)

        if len(opts)==1 and "⚪" in opts[0]:
            st.markdown(f"<span class='na-field'>⚪ {param}</span>", unsafe_allow_html=True)
            answers[key] = opts[0]
            continue

        # Try long key first, then short name (AUDIT MATRIX uses short column names)
        short_key = key.split("|")[-1].strip() if "|" in key else key
        if key in radar_row.index and str(radar_row[key]) not in ["", "nan", "None"]:
            existing = str(radar_row[key])
        elif short_key in radar_row.index and str(radar_row[short_key]) not in ["", "nan", "None"]:
            existing = str(radar_row[short_key])
        else:
            existing = ""
        # Pre-fill: use value from AUDIT MATRIX (which already has last audit data)
        default_idx = opts.index(existing) if existing in opts else 0
        is_mand     = (not is_3d(type_val)) and (key in PA_DTM_MANDATORY_2D)
        label       = f"**{param}**" + (" 🔒" if is_mand else "")

        answers[key] = st.radio(label, options=opts, index=default_idx, key=f"chk_{key}", horizontal=True)

    st.divider()

    pmp_answers = {}  # kept for save_audit signature compatibility
    score_live, issues_live, actions_live = compute_score(answers, type_val, pa_stl)
    s1, s2, s3 = st.columns([1,1,2])
    with s1:
        css = "score-green" if score_live==1.0 else ("score-yellow" if score_live>=0.7 else "score-red")
        st.markdown(f"<div class='{css}'>🩺 {score_live:.0%}</div>", unsafe_allow_html=True)
    with s2:
        st.metric("🔴 Critical", sum(1 for v in answers.values() if "🔴" in str(v)))
        st.metric("⚠️ Warnings", sum(1 for v in answers.values() if "⚠️" in str(v)))
    with s3:
        if issues_live:
            st.markdown("**Issues & Recommended Actions:**")
            for iss, act in zip(issues_live, actions_live):
                st.markdown(f"- **{iss}:** _{act}_")
        else:
            st.success("✅ All checks passed")

    st.divider()

    if not auditor_input.strip():
        st.warning("⚠️ Enter your name before saving.")
    else:
        # Correction mode toggle
        is_correction = st.checkbox(
            "✏️ This is a correction to the last audit — do not count as a new audit",
            value=False,
            help="Check this if you made a mistake and want to overwrite the last audit without adding a new entry to the history."
        )
        if is_correction:
            st.info("ℹ️ Correction mode — this will overwrite the last audit entry for this radar. Total audit count will not change.")

        btn_label = "✏️ Save Correction" if is_correction else "💾 Save Audit to Excel"
        if st.button(btn_label, type="primary", use_container_width=True):
            with st.spinner("Writing to Excel..."):
                ok, msg = save_audit(selected, site_val, type_val, auditor_input.strip(),
                                     audit_date_input, answers, score_live, issues_live, actions_live,
                                     notes=notes_input.strip(), pmp_answers=pmp_answers,
                                     is_correction=is_correction)
            if ok:
                verb = "Correction saved" if is_correction else "Saved"
                st.success(f"✅ {verb}! **{selected}** · Score: **{score_live:.0%}** · {audit_date_input.strftime('%d %b %Y')} · {auditor_input}")
                if not is_correction:
                    st.balloons()
                time.sleep(1)
                st.rerun()
            else:
                st.error(f"❌ {msg}")

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE 2 — HISTORY & TRENDS
# ═══════════════════════════════════════════════════════════════════════════════
elif "History" in page:
    _render_sidebar(page)

    wb      = get_workbook()
    df_hist = get_history(wb)
    df_mat  = get_audit_matrix(wb)

    if df_hist.empty:
        st.markdown("## 📊 Fleet Analytics")
        st.info("📭 No audit history yet. Complete your first audit to see analytics here.")
        st.stop()

    df_hist = df_hist.dropna(subset=["Audit Date"])
    fleet_size    = len(df_mat)
    unique_radars = df_hist["Radar Name"].nunique()
    avg_score     = df_hist["Score Captured"].mean()
    n_perfect     = (df_hist["Score Captured"] == 1.0).sum()
    n_warn        = ((df_hist["Score Captured"] >= 0.7) & (df_hist["Score Captured"] < 1.0)).sum()
    n_crit        = (df_hist["Score Captured"] < 0.7).sum()
    coverage_pct  = unique_radars / fleet_size if fleet_size else 0

    # ── Header ────────────────────────────────────────────────────────────────
    st.markdown('## 📊 Fleet Analytics')
    st.caption(f"Based on {len(df_hist)} audits across {unique_radars} radars · {fleet_size - unique_radars} not yet audited")
    st.divider()

    # ── KPI row ───────────────────────────────────────────────────────────────
    def kpi_card(label, value, sub, color):
        st.markdown(f"""
        <div class="dqp-kpi" style="--kpi-color:{color}">
            <div class="kpi-label">{label}</div>
            <div class="kpi-value">{value}</div>
            <div class="kpi-sub">{sub}</div>
        </div>""", unsafe_allow_html=True)

    k1,k2,k3,k4,k5 = st.columns(5)
    with k1: kpi_card("Fleet Coverage",   f"{coverage_pct:.0%}",
                       f"{unique_radars} of {fleet_size} radars", "#F78F1E")
    with k2: kpi_card("Avg Score",        f"{avg_score:.0%}" if pd.notna(avg_score) else "—",
                       f"{len(df_hist)} total audits", "#3B82F6")
    with k3: kpi_card("🟢 Good",          str(int(n_perfect)),
                       "100% score audits", "#00B050")
    with k4: kpi_card("⚠️ Needs Attention",str(int(n_warn)),
                       "70–99% score audits", "#F59E0B")
    with k5: kpi_card("🔴 Critical",      str(int(n_crit)),
                       "<70% score audits", "#DC2626")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Internal tabs ─────────────────────────────────────────────────────────
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["📈 Trends", "🌎 BU Comparison", "🏆 Site Ranking", "🛰️ Radar Performance", "📋 Audit Log"])

    # ── TAB 1: TRENDS ─────────────────────────────────────────────────────────
    with tab1:
        df_hist["Month"] = df_hist["Audit Date"].dt.to_period("M").dt.to_timestamp()

        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown("#### Score Trend Over Time")
            score_trend = df_hist.groupby("Month")["Score Captured"].mean().reset_index()
            score_trend["Score Captured"] = (score_trend["Score Captured"] * 100).round(1)
            st.line_chart(score_trend.set_index("Month")["Score Captured"],
                          color="#F78F1E", use_container_width=True)
            st.caption("Average score per month across all audits")

        with col_b:
            st.markdown("#### Audit Activity")
            monthly = df_hist.groupby("Month").size().reset_index(name="Audits")
            st.bar_chart(monthly.set_index("Month")["Audits"],
                         color="#393B41", use_container_width=True)
            st.caption("Number of audits completed per month")

        st.markdown("#### Score Distribution")
        dist_cols = st.columns(3)
        total_aud = len(df_hist)
        with dist_cols[0]:
            pct = n_perfect/total_aud if total_aud else 0
            st.markdown(f"""
            <div style="background:#F0FDF4;border-radius:10px;padding:20px;text-align:center">
                <div style="font-size:2.5rem;font-weight:800;color:#00B050">{pct:.0%}</div>
                <div style="font-size:.9rem;font-weight:600;color:#166534">🟢 Perfect (100%)</div>
                <div style="font-size:.8rem;color:#6B7280">{int(n_perfect)} audits</div>
            </div>""", unsafe_allow_html=True)
        with dist_cols[1]:
            pct = n_warn/total_aud if total_aud else 0
            st.markdown(f"""
            <div style="background:#FFFBEB;border-radius:10px;padding:20px;text-align:center">
                <div style="font-size:2.5rem;font-weight:800;color:#D97706">{pct:.0%}</div>
                <div style="font-size:.9rem;font-weight:600;color:#92400E">⚠️ Needs Attention</div>
                <div style="font-size:.8rem;color:#6B7280">{int(n_warn)} audits</div>
            </div>""", unsafe_allow_html=True)
        with dist_cols[2]:
            pct = n_crit/total_aud if total_aud else 0
            st.markdown(f"""
            <div style="background:#FEF2F2;border-radius:10px;padding:20px;text-align:center">
                <div style="font-size:2.5rem;font-weight:800;color:#DC2626">{pct:.0%}</div>
                <div style="font-size:.9rem;font-weight:600;color:#991B1B">🔴 Critical</div>
                <div style="font-size:.8rem;color:#6B7280">{int(n_crit)} audits</div>
            </div>""", unsafe_allow_html=True)

        # Auditor breakdown
        st.markdown("<br>#### Audits by Auditor", unsafe_allow_html=True)
        by_aud = df_hist.groupby("Auditor").agg(
            Audits=("Score Captured","count"),
            Avg_Score=("Score Captured","mean")
        ).reset_index().sort_values("Audits", ascending=False)
        by_aud["Avg Score"] = by_aud["Avg_Score"].apply(lambda x: f"{x:.0%}" if pd.notna(x) else "—")
        st.dataframe(
            by_aud[["Auditor","Audits","Avg Score"]],
            use_container_width=True, hide_index=True
        )

        # ── Top failing checkpoints ────────────────────────────────────────
        st.markdown("<br>#### 🔍 Most Common Issues Across Fleet", unsafe_allow_html=True)
        st.caption("Checkpoints most frequently marked 🔴 Critical or ⚠️ Needs Attention across all audits")

        wb_top   = get_workbook()
        df_top   = get_audit_matrix(wb_top)
        META_TOP = {"Radar Name","Site","Type","Remote Access","Last Audit","Total Audits",
                    "Auditor","SCORE","ISSUES DETECTED","RECOMMENDED ACTION",
                    "HELPER_SCORE","HELPER_DATE","HELPER_STATUS","PA","BU"}
        dqp_top  = [c for c in df_top.columns if c not in META_TOP and c and not str(c).startswith("_")]

        fail_counts = {}
        warn_counts = {}
        for c in dqp_top:
            fail_counts[c] = df_top[c].apply(lambda v: "🔴" in str(v)).sum()
            warn_counts[c] = df_top[c].apply(lambda v: "⚠️" in str(v)).sum()

        # Combine and sort
        all_issues = {c: fail_counts[c] + warn_counts[c] for c in dqp_top if fail_counts[c] + warn_counts[c] > 0}
        top10 = sorted(all_issues.items(), key=lambda x: -x[1])[:10]

        if not top10:
            st.info("No issues found yet — audit some radars first.")
        else:
            max_count = top10[0][1]
            for rank, (checkpoint, total) in enumerate(top10):
                crits = fail_counts[checkpoint]
                warns = warn_counts[checkpoint]
                bar_pct = int(total / max_count * 100)
                crit_pct = int(crits / total * 100) if total > 0 else 0
                warn_pct = 100 - crit_pct
                color = "#DC2626" if crits > warns else "#D97706"

                st.markdown(f"""
                <div style="margin:6px 0;padding:10px 14px;background:#F9FAFB;
                            border-radius:8px;border-left:4px solid {color}">
                    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:4px">
                        <span style="font-weight:700;color:#1F2937;font-size:.9rem">
                            {rank+1}. {checkpoint}
                        </span>
                        <span style="font-size:.8rem;color:#6B7280">
                            🔴 {crits} critical &nbsp;·&nbsp; ⚠️ {warns} attention &nbsp;·&nbsp;
                            <b style="color:{color}">{total} total</b>
                        </span>
                    </div>
                    <div style="background:#E5E7EB;border-radius:4px;height:6px">
                        <div style="background:{color};width:{bar_pct}%;height:6px;border-radius:4px"></div>
                    </div>
                </div>""", unsafe_allow_html=True)
    # ── TAB: BU COMPARISON ───────────────────────────────────────────────────
    with tab2:
        st.markdown("#### BU Performance Comparison")
        if "BU" not in df_mat.columns:
            st.info("BU data not available.")
        else:
            _bu_hist = df_mat[["Radar Name","BU"]].merge(
                df_hist[["Radar Name","Audit Date","Score Captured"]],
                on="Radar Name", how="left"
            )
            _bu_summary = []
            for _bu in sorted(df_mat["BU"].dropna().unique()):
                _sub_mat  = df_mat[df_mat["BU"] == _bu]
                _sub_hist = _bu_hist[_bu_hist["BU"] == _bu].dropna(subset=["Audit Date"])
                _n_radars = len(_sub_mat)
                _n_aud    = _sub_mat["Last Audit"].notna().sum()
                _cov      = _n_aud / _n_radars if _n_radars else 0
                _scores   = _sub_hist["Score Captured"].dropna()
                _avg      = _scores.mean() if len(_scores) else None
                _n_crit   = (_scores < 0.7).sum() if len(_scores) else 0
                _n_warn   = ((_scores >= 0.7) & (_scores < 1.0)).sum() if len(_scores) else 0
                _n_good   = (_scores == 1.0).sum() if len(_scores) else 0
                _bu_summary.append({
                    "BU": _bu, "Radars": _n_radars,
                    "Audited": _n_aud, "Coverage": _cov,
                    "Avg Score": _avg, "Total Audits": len(_sub_hist),
                    "Good": _n_good, "Attention": _n_warn, "Critical": _n_crit
                })
            _bu_df = pd.DataFrame(_bu_summary)
            _cc1, _cc2 = st.columns(2)
            with _cc1:
                st.markdown("**Fleet Coverage by BU**")
                _cov_chart = _bu_df[["BU","Coverage"]].copy()
                _cov_chart["Coverage %"] = (_cov_chart["Coverage"] * 100).round(0)
                st.bar_chart(_cov_chart.set_index("BU")["Coverage %"], use_container_width=True, height=220)
            with _cc2:
                st.markdown("**Avg Score by BU**")
                _sc_chart = _bu_df[["BU","Avg Score"]].copy()
                _sc_chart["Avg Score %"] = (_sc_chart["Avg Score"] * 100).round(0).fillna(0)
                st.bar_chart(_sc_chart.set_index("BU")["Avg Score %"], use_container_width=True, height=220)
            st.divider()
            _disp = _bu_df.copy()
            _disp["Coverage"]  = _disp["Coverage"].apply(lambda x: f"{x:.0%}")
            _disp["Avg Score"] = _disp["Avg Score"].apply(lambda x: f"{x:.0%}" if pd.notna(x) else "—")
            _disp_cols = ["BU","Radars","Audited","Coverage","Avg Score","Total Audits","Good","Attention","Critical"]
            st.dataframe(_disp[[c for c in _disp_cols if c in _disp.columns]], use_container_width=True, hide_index=True)

    with tab3:
        st.markdown("#### Site Performance Ranking")
        st.caption("Ranked by average score across all audits. Sites with more audits are more reliable.")

        # Build site stats
        site_stats = df_hist.groupby("Site").agg(
            Audits=("Score Captured","count"),
            Avg=("Score Captured","mean"),
            Last=("Audit Date","max"),
            Radars=("Radar Name","nunique"),
        ).reset_index().sort_values("Avg", ascending=False)

        # Render as visual cards
        for i, (_, row) in enumerate(site_stats.iterrows()):
            avg    = row["Avg"]
            medal  = ["🥇","🥈","🥉"][i] if i < 3 else f"{i+1}."
            color  = "#00B050" if avg==1.0 else ("#F59E0B" if avg>=0.7 else "#DC2626")
            bg     = "#F0FDF4" if avg==1.0 else ("#FFFBEB" if avg>=0.7 else "#FEF2F2")
            bar_w  = int(avg * 100)
            last_d = pd.to_datetime(row["Last"]).strftime("%d %b %Y") if pd.notna(row["Last"]) else "—"
            n_fleet= len(df_mat[df_mat["Site"]==row["Site"]]) if not df_mat.empty else "?"

            st.markdown(f"""
            <div style="background:{bg};border-radius:10px;padding:14px 18px;margin:6px 0;
                        border-left:5px solid {color}">
                <div style="display:flex;justify-content:space-between;align-items:center">
                    <div>
                        <span style="font-size:1.1rem;font-weight:800;color:#1F2937">
                            {medal} {row['Site']}
                        </span>
                        <span style="font-size:.8rem;color:#6B7280;margin-left:10px">
                            {row['Audits']} audits · {row['Radars']}/{n_fleet} radars · Last: {last_d}
                        </span>
                    </div>
                    <div style="font-size:1.6rem;font-weight:800;color:{color}">{avg:.0%}</div>
                </div>
                <div style="background:#E5E7EB;border-radius:4px;height:6px;margin-top:8px">
                    <div style="background:{color};width:{bar_w}%;height:6px;border-radius:4px"></div>
                </div>
            </div>""", unsafe_allow_html=True)

    # ── TAB 4: RADAR PERFORMANCE ──────────────────────────────────────────────
    with tab4:
        st.markdown("#### Fleet Status & Radar Performance")

        # ── Fleet status table ─────────────────────────────────────────────
        # Get last audit per radar from history
        if not df_hist.empty:
            last_by_radar = (
                df_hist.sort_values("Audit Date", ascending=False)
                .groupby("Radar Name").first()
                .reset_index()[["Radar Name","Audit Date","Score Captured","Auditor","Site"]]
            )
        else:
            last_by_radar = pd.DataFrame(columns=["Radar Name","Audit Date","Score Captured","Auditor","Site"])

        # Merge with full fleet
        fleet = df_mat[["Radar Name","Site","Type","Remote Access"]].copy()
        fleet = fleet.merge(last_by_radar[["Radar Name","Audit Date","Score Captured","Auditor"]],
                            on="Radar Name", how="left")

        def fleet_status(sc):
            if pd.isna(sc): return "⚫ Not Audited"
            if sc == 1.0:   return "🟢 Good"
            if sc >= 0.7:   return "⚠️ Needs Attention"
            return "🔴 Critical"

        fleet["Status"] = fleet["Score Captured"].apply(fleet_status)
        fleet["Score"]  = fleet["Score Captured"].apply(lambda x: f"{x:.0%}" if pd.notna(x) else "—")
        fleet["Days Since"] = fleet["Audit Date"].apply(
            lambda d: (pd.Timestamp.now() - pd.to_datetime(d)).days if pd.notna(d) else None
        )

        # ── Filters ────────────────────────────────────────────────────────
        fa_col, fb_col, fc_col, fd_col = st.columns(4)
        search_r = fa_col.text_input("🔍 Search radar or site", placeholder="SSR126 / Arcelormittal")
        filter_s = fb_col.selectbox("Status",
            ["All", "🟢 Good", "⚠️ Needs Attention", "🔴 Critical", "⚫ Not Audited"])
        filter_t = fc_col.selectbox("Type",
            ["All"] + sorted(fleet["Type"].dropna().unique().tolist()))
        _bu_opts = ["All"] + sorted(fleet["BU"].dropna().unique().tolist()) if "BU" in fleet.columns else ["All"]
        filter_bu = fd_col.selectbox("🌎 BU", _bu_opts)

        # Apply filters
        df_fleet = fleet.copy()
        if search_r:
            s = search_r.upper()
            df_fleet = df_fleet[
                df_fleet["Radar Name"].str.upper().str.contains(s) |
                df_fleet["Site"].str.upper().str.contains(s)
            ]
        if filter_s != "All":
            df_fleet = df_fleet[df_fleet["Status"] == filter_s]
        if filter_t != "All":
            df_fleet = df_fleet[df_fleet["Type"] == filter_t]
        if filter_bu != "All" and "BU" in df_fleet.columns:
            df_fleet = df_fleet[df_fleet["BU"] == filter_bu]

        # Sort: Critical first, then Needs Attention, then Not Audited, then Good
        order = {"🔴 Critical": 0, "⚠️ Needs Attention": 1, "⚫ Not Audited": 2, "🟢 Good": 3}
        df_fleet["_sort"] = df_fleet["Status"].map(order)
        df_fleet = df_fleet.sort_values(["_sort","Score Captured"], ascending=[True, True])

        # ── Mini KPI strip ─────────────────────────────────────────────────
        n_good_f  = (fleet["Status"] == "🟢 Good").sum()
        n_warn_f  = (fleet["Status"] == "⚠️ Needs Attention").sum()
        n_crit_f  = (fleet["Status"] == "🔴 Critical").sum()
        n_na_f    = (fleet["Status"] == "⚫ Not Audited").sum()

        strip = st.columns(4)
        for col, label, val, clr, bg in [
            (strip[0], "🟢 Good",           n_good_f, "#00B050", "#F0FDF4"),
            (strip[1], "⚠️ Needs Attention", n_warn_f, "#D97706", "#FFFBEB"),
            (strip[2], "🔴 Critical",        n_crit_f, "#DC2626", "#FEF2F2"),
            (strip[3], "⚫ Not Audited",     n_na_f,   "#6B7280", "#F9FAFB"),
        ]:
            col.markdown(f"""
            <div style="background:{bg};border-radius:8px;padding:10px 14px;text-align:center;
                        border:1px solid {clr}30">
                <div style="font-size:.75rem;font-weight:700;color:{clr}">{label}</div>
                <div style="font-size:1.8rem;font-weight:800;color:{clr};line-height:1.1">{val}</div>
            </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Fleet table ────────────────────────────────────────────────────
        display_cols = {
            "Radar Name": "Radar",
            "Site":       "Site",
            "Type":       "Type",
            "Status":     "Status",
            "Score":      "Score",
            "Days Since": "Days Since Audit",
            "Auditor":    "Last Auditor",
        }
        df_display = df_fleet[list(display_cols.keys())].rename(columns=display_cols)
        df_display["Days Since Audit"] = df_display["Days Since Audit"].apply(
            lambda x: f"{int(x)}d ago" if pd.notna(x) and x == x else "Never"
        )

        st.dataframe(
            df_display,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Status": st.column_config.TextColumn("Status", width="medium"),
                "Score":  st.column_config.TextColumn("Score",  width="small"),
                "Days Since Audit": st.column_config.TextColumn("Days Since Audit", width="medium"),
            }
        )
        st.caption(f"Showing {len(df_fleet)} of {len(fleet)} radars")

        # ── Export fleet to Excel ──────────────────────────────────────────
        def build_fleet_export(fleet_df, df_hist_full):
            """Build a clean Excel export of all radars with their latest audit data."""
            import io as _io
            output = _io.BytesIO()

            # Get last audit per radar
            if not df_hist_full.empty:
                last = (
                    df_hist_full.sort_values("Audit Date", ascending=False)
                    .groupby("Radar Name").first()
                    .reset_index()[["Radar Name","Audit Date","Score Captured","Auditor"]]
                )
                export = fleet_df.merge(last, on="Radar Name", how="left")

            # Ensure Score Captured column always exists
            if "Score Captured" not in export.columns:
                export["Score Captured"] = None
            if "Audit Date" not in export.columns:
                export["Audit Date"] = None
            if "Auditor" not in export.columns:
                export["Auditor"] = None
            else:
                export = fleet_df.copy()
                export["Audit Date"] = None
                export["Score Captured"] = None
                export["Auditor"] = None

            # Format columns
            export["Score"] = export["Score Captured"].apply(
                lambda x: f"{x:.0%}" if pd.notna(x) else "Not Audited"
            )
            export["Status"] = export["Score Captured"].apply(
                lambda x: "Good" if x==1.0 else ("Needs Attention" if x>=0.7 else "Critical") if pd.notna(x) else "Not Audited"
            )
            export["Days Since Audit"] = export["Audit Date"].apply(
                lambda d: int((pd.Timestamp.now()-pd.to_datetime(d)).days) if pd.notna(d) else None
            )
            export["Last Audit Date"] = export["Audit Date"].apply(
                lambda d: pd.to_datetime(d).strftime("%d %b %Y") if pd.notna(d) else "Never"
            )

            cols = ["Radar Name","Site","BU","Type","Remote Access","Status","Score",
                    "Last Audit Date","Days Since Audit","Auditor"]
            export = export[[c for c in cols if c in export.columns]]

            with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                export.to_excel(writer, sheet_name="Fleet DQP Status", index=False)
                wb_x = writer.book
                ws_x = writer.sheets["Fleet DQP Status"]

                # Formats
                hdr_fmt = wb_x.add_format({
                    "bold": True, "bg_color": "#393B41", "font_color": "#FFFFFF",
                    "border": 1, "align": "center"
                })
                green_fmt  = wb_x.add_format({"bg_color": "#C6EFCE", "font_color": "#276221"})
                yellow_fmt = wb_x.add_format({"bg_color": "#FFEB9C", "font_color": "#9C6500"})
                red_fmt    = wb_x.add_format({"bg_color": "#FFC7CE", "font_color": "#9C0006"})
                grey_fmt   = wb_x.add_format({"bg_color": "#F2F2F2", "font_color": "#666666"})

                # Header row
                for col_num, col_name in enumerate(export.columns):
                    ws_x.write(0, col_num, col_name, hdr_fmt)

                # Column widths
                widths = [12, 20, 8, 14, 16, 8, 16, 16, 16]
                for i, w in enumerate(widths[:len(export.columns)]):
                    ws_x.set_column(i, i, w)

                # Color status column
                status_col = list(export.columns).index("Status") if "Status" in export.columns else None
                if status_col is not None:
                    for row_num, status in enumerate(export["Status"], 1):
                        fmt = (green_fmt  if status == "Good" else
                               yellow_fmt if status == "Needs Attention" else
                               red_fmt    if status == "Critical" else grey_fmt)
                        ws_x.write(row_num, status_col, status, fmt)

            return output.getvalue()

        col_exp1, col_exp2 = st.columns([3, 1])
        with col_exp2:
            export_bytes = build_fleet_export(fleet, df_hist)
            st.download_button(
                label="⬇️ Export to Excel",
                data=export_bytes,
                file_name=f"Fleet_DQP_Status_{date.today().strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        st.divider()
        st.markdown("#### 🔎 Radar Audit History")
        st.caption("Select a radar to see its full audit history and score evolution")

        all_radars = sorted(df_mat["Radar Name"].dropna().tolist())
        radar_labels = []
        for r in all_radars:
            t = df_mat.loc[df_mat["Radar Name"]==r, "Type"].values
            t = str(t[0]) if len(t) > 0 else ""
            radar_labels.append(f"{r} · {t}" if t else r)

        label_to_name = {lbl: name for lbl, name in zip(radar_labels, all_radars)}

        sel_label = st.selectbox("Select Radar", radar_labels, key="history_radar_sel")
        sel_radar = label_to_name[sel_label]
        radar_info = df_mat[df_mat["Radar Name"] == sel_radar].iloc[0] if not df_mat.empty else None

        radar_audits = df_hist[df_hist["Radar Name"] == sel_radar].sort_values("Audit Date")

        if radar_info is not None:
            ri1, ri2, ri3 = st.columns(3)
            ri1.markdown(f"**Site:** {radar_info.get('Site','—')}")
            ri2.markdown(f"**Type:** `{radar_info.get('Type','—')}`")
            ri3.markdown(f"**Access:** {radar_info.get('Remote Access','—')}")

        if radar_audits.empty:
            st.info(f"⚪ No audits recorded for {sel_radar} yet.")
        else:
            # Score trend for this radar
            n_aud = len(radar_audits)
            last_sc = radar_audits.iloc[-1]["Score Captured"]
            last_dt = radar_audits.iloc[-1]["Audit Date"]
            avg_sc  = radar_audits["Score Captured"].mean()

            m1, m2, m3 = st.columns(3)
            sc_color = "#00B050" if last_sc==1.0 else ("#D97706" if last_sc>=0.7 else "#DC2626") if pd.notna(last_sc) else "#6B7280"
            m1.markdown(f"""<div style="background:#F5F6F8;border-radius:8px;padding:12px 16px;
                text-align:center;border-left:4px solid {sc_color}">
                <div style="font-size:.75rem;font-weight:700;color:{sc_color};text-transform:uppercase">Latest Score</div>
                <div style="font-size:2rem;font-weight:800;color:{sc_color}">{f"{last_sc:.0%}" if pd.notna(last_sc) else "—"}</div>
                <div style="font-size:.75rem;color:#6B7280">{pd.to_datetime(last_dt).strftime("%d %b %Y") if pd.notna(last_dt) else "—"}</div>
            </div>""", unsafe_allow_html=True)
            m2.markdown(f"""<div style="background:#F5F6F8;border-radius:8px;padding:12px 16px;text-align:center">
                <div style="font-size:.75rem;font-weight:700;color:#3B82F6;text-transform:uppercase">Avg Score</div>
                <div style="font-size:2rem;font-weight:800;color:#3B82F6">{f"{avg_sc:.0%}" if pd.notna(avg_sc) else "—"}</div>
                <div style="font-size:.75rem;color:#6B7280">across {n_aud} audits</div>
            </div>""", unsafe_allow_html=True)
            m3.markdown(f"""<div style="background:#F5F6F8;border-radius:8px;padding:12px 16px;text-align:center">
                <div style="font-size:.75rem;font-weight:700;color:#393B41;text-transform:uppercase">Total Audits</div>
                <div style="font-size:2rem;font-weight:800;color:#393B41">{n_aud}</div>
                <div style="font-size:.75rem;color:#6B7280">since first audit</div>
            </div>""", unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # Score trend chart
            if n_aud > 1:
                st.markdown("**Score History**")
                trend = radar_audits.copy()
                trend["Score %"] = (trend["Score Captured"] * 100).round(1)
                trend["Date"]    = trend["Audit Date"].dt.strftime("%d %b %Y")
                st.line_chart(
                    trend.set_index("Date")["Score %"],
                    color="#F78F1E", use_container_width=True, height=180
                )

            # Full audit log for this radar with comparison
            st.markdown("**Audit History**")
            log = radar_audits.sort_values("Audit Date", ascending=False).copy().reset_index(drop=True)
            log["Date"]   = log["Audit Date"].dt.strftime("%d %b %Y")
            log["Score"]  = log["Score Captured"].apply(lambda x: f"{x:.0%}" if pd.notna(x) else "—")
            log["Status"] = log["Score Captured"].apply(
                lambda x: "🟢 Good" if x==1.0 else ("⚠️ Needs Attention" if x>=0.7 else "🔴 Critical") if pd.notna(x) else "—"
            )

            # Add delta vs previous audit
            def score_delta(i, log_df):
                if i >= len(log_df) - 1: return "—"
                curr = log_df.loc[i, "Score Captured"]
                prev = log_df.loc[i+1, "Score Captured"]
                if pd.isna(curr) or pd.isna(prev): return "—"
                diff = curr - prev
                if abs(diff) < 0.01: return "➡️ Same"
                return f"↑ +{diff:.0%}" if diff > 0 else f"↓ {diff:.0%}"

            log["vs Previous"] = [score_delta(i, log) for i in range(len(log))]

            # Show notes if present
            has_notes = "Notes" in log.columns and log["Notes"].notna().any()
            show_cols = ["Date","Score","Status","vs Previous","Auditor"]
            if has_notes:
                show_cols.append("Notes")

            st.dataframe(
                log[show_cols],
                use_container_width=True, hide_index=True,
                column_config={
                    "Score":       st.column_config.TextColumn("Score",       width="small"),
                    "Status":      st.column_config.TextColumn("Status",      width="medium"),
                    "Date":        st.column_config.TextColumn("Date",        width="small"),
                    "vs Previous": st.column_config.TextColumn("vs Previous", width="medium"),
                    "Notes":       st.column_config.TextColumn("Notes",       width="large"),
                }
            )

            # If 2+ audits, show what changed between last two
            if len(log) >= 2:
                curr_sc = log.loc[0, "Score Captured"]
                prev_sc = log.loc[1, "Score Captured"]
                if pd.notna(curr_sc) and pd.notna(prev_sc):
                    diff = curr_sc - prev_sc
                    if abs(diff) >= 0.01:
                        diff_clr = "#00B050" if diff > 0 else "#DC2626"
                        diff_icon = "↑" if diff > 0 else "↓"
                        diff_txt  = "improved" if diff > 0 else "declined"
                        st.markdown(f"""
                        <div style="background:{'#F0FDF4' if diff>0 else '#FEF2F2'};border-radius:8px;
                                    padding:10px 16px;margin-top:8px;border-left:4px solid {diff_clr}">
                            <span style="font-weight:700;color:{diff_clr}">{diff_icon} {abs(diff):.0%} {diff_txt}</span>
                            <span style="color:#6B7280;font-size:.85rem;margin-left:8px">
                                vs previous audit ({log.loc[1,'Date']})
                            </span>
                        </div>""", unsafe_allow_html=True)

    # ── TAB 5: AUDIT LOG ──────────────────────────────────────────────────────
    with tab5:
        st.markdown("#### Full Audit Log")

        f1, f2, f3 = st.columns(3)
        fa  = f1.selectbox("Auditor", ["All"] + sorted(df_hist["Auditor"].dropna().unique().tolist()))
        fs  = f2.selectbox("Site",    ["All"] + sorted(df_hist["Site"].dropna().unique().tolist()))
        fst = f3.selectbox("Status",  ["All", "🟢 Perfect", "⚠️ Needs Attention", "🔴 Critical"])

        df_show = df_hist.copy()
        if fa != "All": df_show = df_show[df_show["Auditor"] == fa]
        if fs != "All": df_show = df_show[df_show["Site"] == fs]
        if "Perfect"   in fst: df_show = df_show[df_show["Score Captured"] == 1.0]
        elif "Needs"   in fst: df_show = df_show[(df_show["Score Captured"] >= 0.7) & (df_show["Score Captured"] < 1.0)]
        elif "Critical" in fst: df_show = df_show[df_show["Score Captured"] < 0.7]

        df_show = df_show.sort_values("Audit Date", ascending=False).copy()

        # Add status column
        def status_label(sc):
            if pd.isna(sc): return "—"
            if sc == 1.0:   return "🟢 Perfect"
            if sc >= 0.7:   return "⚠️ Needs Attention"
            return "🔴 Critical"

        df_show["Status"] = df_show["Score Captured"].apply(status_label)
        df_show["Score"]  = df_show["Score Captured"].apply(lambda x: f"{x:.0%}" if pd.notna(x) else "—")
        df_show["Date"]   = df_show["Audit Date"].dt.strftime("%d %b %Y")

        _log_cols = ["Date","Radar Name","Site","Auditor","Score","Status"]
        if "Notes" in df_show.columns:
            df_show["Notes"] = df_show["Notes"].fillna("")
            _log_cols.append("Notes")
        st.dataframe(
            df_show[_log_cols],
            use_container_width=True, hide_index=True,
            column_config={
                "Score":  st.column_config.TextColumn("Score",  width="small"),
                "Status": st.column_config.TextColumn("Status", width="medium"),
                "Date":   st.column_config.TextColumn("Date",   width="small"),
                "Notes":  st.column_config.TextColumn("Notes",  width="large"),
            }
        )
        st.caption(f"{len(df_show)} of {len(df_hist)} records")




# ═══════════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
# PAGE 3 — EXPIRATIONS
# ═══════════════════════════════════════════════════════════════════════════════
elif "Expir" in page:
    _render_sidebar(page)
    st.markdown('## ⏰ Audit Expirations')

    wb      = get_workbook()
    df_mat  = get_audit_matrix(wb)
    df_hist = get_history(wb)

    today         = datetime.now()
    EXPIRY_DAYS   = 180
    WARNING_DAYS  = 30

    # Filters row
    fc1, fc2, fc3, fc4 = st.columns([2,1,1,1])
    exp_search = fc1.text_input("🔍 Search radar or site", placeholder="SSR171 / Sishen Mine", label_visibility="collapsed")
    _exp_bus   = ["All BUs"] + sorted(df_mat["BU"].dropna().unique().tolist()) if "BU" in df_mat.columns else ["All BUs"]
    exp_bu     = fc2.selectbox("BU", _exp_bus, label_visibility="collapsed")
    _exp_types = ["All Types"] + sorted(df_mat["Type"].dropna().unique().tolist())
    exp_type   = fc3.selectbox("Type", _exp_types, label_visibility="collapsed")
    exp_status = fc4.selectbox("Status", ["All","🔴 Expired","⬛ Never Audited","⚠️ Expiring Soon","🟢 Up to Date"], label_visibility="collapsed")

    # Build last audit — HISTORY_LOG only, only entries with a real Score ────────
    if not df_hist.empty:
        _hist_scored = df_hist[df_hist["Score Captured"].notna()].dropna(subset=["Audit Date"]).copy()
        if not _hist_scored.empty:
            last_by_radar = (
                _hist_scored
                .sort_values("Audit Date", ascending=False)
                .groupby("Radar Name").first().reset_index()
                [["Radar Name","Audit Date","Score Captured","Auditor"]]
            )
        else:
            last_by_radar = pd.DataFrame(columns=["Radar Name","Audit Date","Score Captured","Auditor"])
    else:
        last_by_radar = pd.DataFrame(columns=["Radar Name","Audit Date","Score Captured","Auditor"])

    _hist_count = len(last_by_radar)

    # Only take identity fields from AUDIT MATRIX — never Last Audit (may have demo dates)
    keep_cols = ["Radar Name","Site","Type","Remote Access"]
    if "BU" in df_mat.columns:
        keep_cols.append("BU")
    _df_status_base = df_mat[[c for c in keep_cols if c in df_mat.columns]].copy()
    # Ensure no Last Audit leaks from AUDIT MATRIX
    for _drop_col in ["Last Audit","Total Audits","Auditor","SCORE","HELPER_SCORE"]:
        if _drop_col in _df_status_base.columns:
            _df_status_base = _df_status_base.drop(columns=[_drop_col])
    df_status = _df_status_base.merge(last_by_radar, on="Radar Name", how="left")
    df_status["Days Since"] = df_status["Audit Date"].apply(
        lambda d: (today - pd.to_datetime(d)).days if pd.notna(d) else None)
    df_status["Expires In"] = df_status["Days Since"].apply(
        lambda d: EXPIRY_DAYS - d if d is not None else None)

    def _exp_status(row):
        ds = row["Days Since"]
        if ds is None:                         return "⬛ Never Audited"
        if ds > EXPIRY_DAYS:                   return "🔴 Expired"
        if ds >= EXPIRY_DAYS - WARNING_DAYS:   return "⚠️ Expiring Soon"
        return "🟢 Up to Date"

    df_status["Status"] = df_status.apply(_exp_status, axis=1)

    # Apply filters
    df_filt = df_status.copy()
    if exp_search:
        s = exp_search.upper()
        df_filt = df_filt[df_filt["Radar Name"].str.upper().str.contains(s, na=False) |
                          df_filt["Site"].str.upper().str.contains(s, na=False)]
    if exp_bu != "All BUs" and "BU" in df_filt.columns:
        df_filt = df_filt[df_filt["BU"] == exp_bu]
    if exp_type != "All Types":
        df_filt = df_filt[df_filt["Type"] == exp_type]
    if exp_status != "All":
        df_filt = df_filt[df_filt["Status"] == exp_status]

    # KPI row — reflects BU/Type filter
    _base = df_status.copy()
    if exp_bu != "All BUs" and "BU" in _base.columns:
        _base = _base[_base["BU"] == exp_bu]
    if exp_type != "All Types":
        _base = _base[_base["Type"] == exp_type]

    k1,k2,k3,k4 = st.columns(4)
    k1.metric("🔴 Expired",      (_base["Status"]=="🔴 Expired").sum(),      help=">180 days")
    k2.metric("⬛ Never Audited", (_base["Status"]=="⬛ Never Audited").sum(), help="No audit on record")
    k3.metric("⚠️ Expiring Soon", (_base["Status"]=="⚠️ Expiring Soon").sum(),help="Within 30 days")
    k4.metric("🟢 Up to Date",   (_base["Status"]=="🟢 Up to Date").sum(),   help="Within 150 days")

    st.caption(f"Showing {len(df_filt)} of {len(df_status)} radars · {_hist_count} with real audit history")
    st.divider()

    def render_exp_table(df_sub, label, icon):
        if df_sub.empty:
            with st.expander(f"{icon} {label} — 0 radars", expanded=False):
                st.success(f"✅ No radars in this category.")
            return
        with st.expander(f"{icon} {label} — {len(df_sub)} radars", expanded=True):
            rows = []
            for _, r in df_sub.sort_values("Days Since", ascending=False, na_position="last").iterrows():
                ds  = r["Days Since"]
                ei  = r["Expires In"]
                la  = pd.to_datetime(r["Audit Date"]).strftime("%d %b %Y") if pd.notna(r.get("Audit Date")) else "—"
                sc  = f"{r['Score Captured']:.0%}" if pd.notna(r.get("Score Captured")) else "—"
                row_d = {
                    "Radar":      str(r["Radar Name"]),
                    "Site":       str(r["Site"]),
                    "Type":       str(r["Type"]),
                    "Access":     str(r.get("Remote Access","") or ""),
                    "Last Audit": la,
                    "Days Since": f"{int(ds)}d" if ds is not None and pd.notna(ds) else "—",
                    "Expires In": f"{int(ei)}d" if ei is not None and pd.notna(ei) else "—",
                    "Last Score": sc,
                }
                if "BU" in r.index:
                    row_d["BU"] = str(r.get("BU","") or "")
                rows.append(row_d)
            df_out = pd.DataFrame(rows)
            if "BU" in df_out.columns:
                cols = ["BU","Radar","Site","Type","Access","Last Audit","Days Since","Expires In","Last Score"]
                df_out = df_out[[c for c in cols if c in df_out.columns]]
            st.dataframe(df_out, use_container_width=True, hide_index=True)

    if exp_status in ("All","🔴 Expired"):
        render_exp_table(df_filt[df_filt["Status"]=="🔴 Expired"],       "Expired (>180 days)",      "🔴")
    if exp_status in ("All","⚠️ Expiring Soon"):
        render_exp_table(df_filt[df_filt["Status"]=="⚠️ Expiring Soon"], "Expiring Within 30 Days",  "⚠️")
    if exp_status in ("All","⬛ Never Audited"):
        render_exp_table(df_filt[df_filt["Status"]=="⬛ Never Audited"],  "Never Audited",            "⚫")
    if exp_status in ("All","🟢 Up to Date"):
        render_exp_table(df_filt[df_filt["Status"]=="🟢 Up to Date"],    "Up to Date",               "🟢")


# PAGE 4 — CLIENT PDF REPORT
# ═══════════════════════════════════════════════════════════════════════════════
elif "PDF" in page:
    _render_sidebar(page)
    st.markdown("## 📄 Client PDF Report")
    st.caption("Generate a professional Proactive Data Quality Review for your client.")

    wb       = get_workbook()
    df_mat   = get_audit_matrix(wb)
    df_hist  = get_history(wb)

    # Site selector
    sites = sorted(df_mat["Site"].dropna().unique().tolist())
    # BU filter for PDF page
    _pdf_bus = ["All BUs"] + sorted(set(
        v for v in df_mat["BU"].dropna().unique().tolist()
        if v and v != "Other"
    )) if "BU" in df_mat.columns else ["All BUs"]

    fc1, fc2, fc3 = st.columns([1, 2, 1])
    with fc1:
        pdf_bu_filter = st.selectbox("🌎 Business Unit", _pdf_bus)
    with fc2:
        if pdf_bu_filter != "All BUs":
            _bu_sites = sorted(df_mat[df_mat["BU"] == pdf_bu_filter]["Site"].dropna().unique().tolist())
        else:
            _bu_sites = sites
        site_sel = st.selectbox("📍 Select Site", _bu_sites)
    with fc3:
        auditor_pdf = st.text_input("👤 Reviewed by", placeholder="Your name")

    # Site preview
    site_radars = df_mat[df_mat["Site"].str.strip().str.lower() == site_sel.strip().lower()]
    n = len(site_radars)
    st.markdown(f"**{n} radar{'s' if n!=1 else ''}** at {site_sel}")

    # Quick score preview — use same column detection as PDF function
    _META = {'Radar Name','Site','Type','Remote Access','Last Audit','Total Audits','Auditor',
             'SCORE','ISSUES DETECTED','RECOMMENDED ACTION','HELPER_SCORE','HELPER_DATE',
             'HELPER_STATUS','PA','_score'}
    if not site_radars.empty:
        dqp_cols_prev = [c for c in site_radars.columns if c not in _META and c and not str(c).startswith('_')]
        _SH_PREV = {"Data Availability", "SSR Type & Scan Mode", "Signal Strength", "Return Signal", "Wall %"}
        scores = []
        for _, row in site_radars.iterrows():
            vals  = [str(row.get(c,"") or "") for c in dqp_cols_prev]
            reds  = sum(1 for v in vals if "🔴" in v)
            grs   = sum(1 for v in vals if "🟢" in v)
            nas   = sum(1 for v in vals if "⚪" in v)
            total = len(dqp_cols_prev) - nas
            if grs + reds + nas == 0:
                scores.append(float('nan'))
            elif any("🔴" in str(row.get(c,"") or "") for c in _SH_PREV):
                scores.append(0.0)
            else:
                scores.append(grs / total if total > 0 else 1.0)

        preview_data = {
            "Radar":  site_radars["Radar Name"].tolist(),
            "Type":   site_radars["Type"].tolist(),
            "Score":  [f"{s:.0%}" if not pd.isna(s) else "—" for s in scores],
            "Status": ["⚫ Not Audited" if pd.isna(s) else ("🟢 Good" if s==1.0 else ("⚠️ Warning" if s>=0.7 else "🔴 Critical")) for s in scores]
        }
        st.dataframe(preview_data, use_container_width=True, hide_index=True)

    st.divider()

    if not REPORTLAB_OK:
        st.error("❌ reportlab not installed. Add `reportlab` to requirements.txt and redeploy.")
    elif not auditor_pdf.strip():
        st.warning("⚠️ Enter your name before generating the report.")
    else:
        if st.button("📄 Generate PDF Report", type="primary", use_container_width=True):
            with st.spinner("Generating PDF..."):
                pdf_bytes = generate_client_pdf(site_sel, df_mat, auditor_pdf.strip(), logo_path="/home/claude/gp_logo.png")
            fname = f"DQP_Review_{site_sel.replace(' ','_')}_{date.today().strftime('%Y-%m-%d')}.pdf"
            st.download_button(
                label=f"⬇️ Download {fname}",
                data=pdf_bytes,
                file_name=fname,
                mime="application/pdf",
                use_container_width=True
            )
            st.success(f"✅ Report ready — {site_sel} · {len(site_radars)} radars · {date.today().strftime('%d %b %Y')}")


# ═══════════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE 5 — FLEET MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════
elif "Fleet Management" in page:
    _render_sidebar(page)
    st.markdown('## ⚙️ Fleet Management')
    st.caption("Manage radars, import fleet updates, and backup your data.")

    wb     = get_workbook()
    df_mat = get_audit_matrix(wb)

    tab_fleet, tab_add, tab_edit, tab_delete, tab_import, tab_admin = st.tabs([
        "📡 Fleet Overview", "➕ Add Radar", "✏️ Edit Radar", "🗑️ Remove Radar", "📥 Import / Backup", "🔧 Admin"
    ])

    # ── TAB: FLEET OVERVIEW ───────────────────────────────────────────────────
    with tab_fleet:
        st.markdown("#### Fleet Overview")
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Total Radars",   len(df_mat))
        m2.metric("Sites",          df_mat["Site"].nunique())
        m3.metric("Business Units", df_mat["BU"].nunique() if "BU" in df_mat.columns else "—")
        m4.metric("Audited",        df_mat["Last Audit"].notna().sum(), help="At least one audit on record")
        st.divider()
        ov1,ov2,ov3 = st.columns([2,1,1])
        ov_search = ov1.text_input("🔍 Search", placeholder="Radar or site", label_visibility="collapsed")
        _ov_bus   = ["All BUs"] + sorted(df_mat["BU"].dropna().unique().tolist()) if "BU" in df_mat.columns else ["All BUs"]
        ov_bu     = ov2.selectbox("BU",   _ov_bus,   label_visibility="collapsed", key="ov_bu")
        _ov_types = ["All Types"] + sorted(df_mat["Type"].dropna().unique().tolist())
        ov_type   = ov3.selectbox("Type", _ov_types, label_visibility="collapsed", key="ov_type")
        df_ov = df_mat.copy()
        if ov_search:
            s = ov_search.upper()
            df_ov = df_ov[df_ov["Radar Name"].str.upper().str.contains(s,na=False)|
                          df_ov["Site"].str.upper().str.contains(s,na=False)]
        if ov_bu   != "All BUs"   and "BU"   in df_ov.columns: df_ov = df_ov[df_ov["BU"]   == ov_bu]
        if ov_type != "All Types" and "Type" in df_ov.columns: df_ov = df_ov[df_ov["Type"] == ov_type]
        st.caption(f"{len(df_ov)} radars shown")

        # BU summary strip
        if "BU" in df_ov.columns:
            _bu_counts = df_ov["BU"].value_counts()
            _bu_strip  = "  ·  ".join(f"**{bu}** {cnt}" for bu, cnt in _bu_counts.items())
            st.caption(_bu_strip)

        disp_cols = [c for c in ["Radar Name","Site","BU","Type","Remote Access","Last Audit","Total Audits"] if c in df_ov.columns]
        # Replace NaN/None with — for display
        _df_display = df_ov[disp_cols].reset_index(drop=True).copy()
        for _col in ["Last Audit","Total Audits"]:
            if _col in _df_display.columns:
                _df_display[_col] = _df_display[_col].apply(
                    lambda x: "—" if x is None or (isinstance(x, float) and pd.isna(x)) or str(x) in ["nan","None","NaT"] else x)
        st.dataframe(
            _df_display,
            use_container_width=True, hide_index=True,
            column_config={
                "Last Audit":    st.column_config.TextColumn("Last Audit"),
                "Total Audits":  st.column_config.TextColumn("Audits"),
                "Remote Access": st.column_config.TextColumn("Access"),
            }
        )

        # Quick stats below table
        _no_access = df_ov["Remote Access"].isna().sum() + (df_ov["Remote Access"] == "").sum() if "Remote Access" in df_ov.columns else 0
        _no_audit  = df_ov["Last Audit"].isna().sum() if "Last Audit" in df_ov.columns else 0
        if _no_access > 0 or _no_audit > 0:
            _warn_parts = []
            if _no_access > 0: _warn_parts.append(f"⚠️ {_no_access} missing Remote Access")
            if _no_audit  > 0: _warn_parts.append(f"⚫ {_no_audit} never audited")
            st.warning("  ·  ".join(_warn_parts))

    # ── TAB: ADD RADAR ────────────────────────────────────────────────────────
    with tab_add:
        st.markdown("#### Add a new radar to the fleet")
        st.info("The new radar will appear immediately in the Audit Radar page.")
        col1,col2 = st.columns(2)
        with col1:
            new_name  = st.text_input("Radar Name *", placeholder="e.g. SSR999")
            new_site  = st.text_input("Site *",        placeholder="e.g. Arcelormittal")
            new_type  = st.selectbox("Type *", ["XT","FX","SOM","SAR-X"])
        with col2:
            _ebus = sorted(set(v for v in df_mat["BU"].dropna().unique() if v and v!="Other")) if "BU" in df_mat.columns else []
            _bu_choices = _ebus + (["➕ New BU (type below)"] if _ebus else [])
            if _bu_choices:
                _bu_sel = st.selectbox("Business Unit *", [""]+_bu_choices,
                    format_func=lambda x: "Select or add BU..." if x=="" else x, key="new_bu_sel")
                new_bu = st.text_input("Type new BU", placeholder="e.g. GPSA", key="new_bu_custom")                     if _bu_sel in ("➕ New BU (type below)","") else _bu_sel
            else:
                new_bu = st.text_input("Business Unit *", placeholder="e.g. GPNA")
            st.markdown("**Remote Access**")
            _ra_std_opts = ["GSS","VPN","Bomgar","TeamViewer","Anydesk","Customer","Other..."]
            _ra_sel = st.radio("", _ra_std_opts, horizontal=True, key="new_access_radio")
            if _ra_sel == "Other...":
                new_access = st.text_input("Specify access method", placeholder="e.g. RDP, Citrix, VNC", key="new_access_custom")
                if not new_access.strip():
                    new_access = "Other"
            else:
                new_access = _ra_sel
        new_pa = "Yes" if new_type in ["FX","SOM"] else "N/A"
        if st.button("➕ Add Radar", type="primary"):
            if not new_name.strip() or not new_site.strip() or not new_bu.strip():
                st.error("❌ Radar Name, Site and BU are required.")
            elif new_name.strip() in df_mat["Radar Name"].astype(str).tolist():
                st.error(f"❌ **{new_name.strip()}** already exists.")
            else:
                try:
                    ws_am   = wb["AUDIT MATRIX"]
                    headers = [c.value for c in ws_am[1]]
                    next_r  = next((r for r in range(2, ws_am.max_row+2) if ws_am.cell(r,1).value is None), ws_am.max_row+1)
                    col_map = {h:i+1 for i,h in enumerate(headers) if h}
                    ws_am.cell(next_r, col_map.get("Radar Name",1)).value = new_name.strip()
                    ws_am.cell(next_r, col_map.get("Site",2)).value       = new_site.strip()
                    ws_am.cell(next_r, col_map.get("Type",3)).value       = new_type
                    if "Remote Access" in col_map: ws_am.cell(next_r, col_map["Remote Access"]).value = new_access
                    if "PA"            in col_map: ws_am.cell(next_r, col_map["PA"]).value            = new_pa
                    bu_col = ensure_bu_column(ws_am)
                    ws_am.cell(next_r, bu_col).value = new_bu.strip().upper()
                    save_workbook(wb); load_wb_bytes.clear()
                    st.success(f"✅ **{new_name.strip()}** added — {new_site.strip()} · {new_type} · {new_bu.strip().upper()}")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ {e}")

    # ── TAB: EDIT RADAR ───────────────────────────────────────────────────────
    with tab_edit:
        st.markdown("#### Edit radar details")
        st.info("💡 To correct audit answers — go to Audit Radar, select the radar, fix values and Save again.")
        st.caption("Only edits Site, Type, Remote Access and BU. Audit data is not affected.")
        _edit_bu_opts = ["All BUs"] + sorted(df_mat["BU"].dropna().unique().tolist()) if "BU" in df_mat.columns else ["All BUs"]
        ed_bu_filt    = st.selectbox("Filter by BU", _edit_bu_opts, key="edit_bu_filt")
        _edit_df      = df_mat if ed_bu_filt=="All BUs" else df_mat[df_mat["BU"]==ed_bu_filt]
        _edit_list    = sorted(_edit_df["Radar Name"].dropna().astype(str).tolist())
        if not _edit_list:
            st.warning("No radars in this BU.")
        else:
            def _rlabel_e(r):
                rows=df_mat[df_mat["Radar Name"]==r]; t=str(rows["Type"].values[0] if len(rows)>0 else ""); s=str(rows["Site"].values[0] if len(rows)>0 else "")
                return " · ".join(p for p in [r,t,s] if p and p!="nan")
            sel_lbl_e = st.selectbox("Select Radar to Edit", [_rlabel_e(r) for r in _edit_list], key="edit_sel")
            sel_r_e   = _edit_list[[_rlabel_e(r) for r in _edit_list].index(sel_lbl_e)]
            row_e     = df_mat[df_mat["Radar Name"]==sel_r_e].iloc[0]
            cur_site   = str(row_e["Site"])   if pd.notna(row_e.get("Site"))         else ""
            cur_type   = str(row_e["Type"])   if pd.notna(row_e.get("Type"))         else "XT"
            cur_access = str(row_e["Remote Access"]) if pd.notna(row_e.get("Remote Access")) else "GSS"
            _bfe = str(row_e.get("BU","") or "") if "BU" in row_e.index else ""
            cur_bu = _bfe if _bfe and _bfe!="nan" else SITE_BU.get(cur_site.strip(),"")
            ec1,ec2 = st.columns(2)
            with ec1:
                edit_site  = st.text_input("Site", value=cur_site, key=f"e_site_{sel_r_e}")
                _type_opts = ["XT","FX","SOM","SAR-X"]
                edit_type  = st.selectbox("Type", _type_opts, index=_type_opts.index(cur_type) if cur_type in _type_opts else 0, key=f"e_type_{sel_r_e}")
                _ebus2 = sorted(set(v for v in df_mat["BU"].dropna().unique() if v and v!="Other")) if "BU" in df_mat.columns else []
                _ebu_choices = _ebus2 + (["➕ New BU (type below)"] if _ebus2 else [])
                if _ebu_choices:
                    _ebu_sel = st.selectbox("Business Unit", [""]+_ebu_choices,
                        index=([""]+_ebu_choices).index(cur_bu) if cur_bu in _ebu_choices else 0,
                        format_func=lambda x:"Select BU..." if x=="" else x, key=f"e_bu_sel_{sel_r_e}")
                    edit_bu = st.text_input("Type new BU", value=cur_bu if cur_bu not in _ebu_choices else "",
                                            placeholder="e.g. GPSA", key=f"e_bu_{sel_r_e}")                         if _ebu_sel in ("➕ New BU (type below)","") else _ebu_sel
                else:
                    edit_bu = st.text_input("Business Unit", value=cur_bu, key=f"e_bu_{sel_r_e}")
            with ec2:
                _acc_opts = ["GSS","VPN","Bomgar","TeamViewer","Anydesk","Customer","Other..."]
                st.markdown("**Remote Access**")
                _cur_ra_idx = _acc_opts.index(cur_access) if cur_access in _acc_opts else len(_acc_opts)-1
                _ra_edit_sel = st.radio("", _acc_opts, horizontal=True,
                                        index=_cur_ra_idx,
                                        key=f"e_access_{sel_r_e}")
                if _ra_edit_sel == "Other...":
                    edit_access = st.text_input("Specify access method",
                                                value=cur_access if cur_access not in _acc_opts[:-1] else "",
                                                placeholder="e.g. RDP, Citrix, VNC",
                                                key=f"e_access_custom_{sel_r_e}")
                    if not edit_access.strip():
                        edit_access = "Other"
                else:
                    edit_access = _ra_edit_sel
            edit_pa = "Yes" if edit_type in ["FX","SOM"] else "N/A"
            if st.button("💾 Save Changes", type="primary", key="save_edit"):
                try:
                    ws_am        = wb["AUDIT MATRIX"]
                    bu_col_idx_e = ensure_bu_column(ws_am)
                    headers      = [c.value for c in ws_am[1]]
                    col_map      = {h:i+1 for i,h in enumerate(headers) if h}
                    target       = next((r[0].row for r in ws_am.iter_rows(min_row=2) if str(r[0].value)==sel_r_e), None)
                    if target:
                        if "Site"          in col_map: ws_am.cell(target,col_map["Site"]).value          = edit_site.strip()
                        if "Type"          in col_map: ws_am.cell(target,col_map["Type"]).value          = edit_type
                        if "Remote Access" in col_map: ws_am.cell(target,col_map["Remote Access"]).value = edit_access
                        if "PA"            in col_map: ws_am.cell(target,col_map["PA"]).value            = edit_pa
                        if edit_bu.strip():            ws_am.cell(target,bu_col_idx_e).value             = edit_bu.strip().upper()
                        save_workbook(wb); load_wb_bytes.clear()
                        st.success(f"✅ **{sel_r_e}** updated."); st.rerun()
                    else:
                        st.error("Radar not found.")
                except Exception as e:
                    st.error(f"❌ {e}")

    # ── TAB: REMOVE RADAR ─────────────────────────────────────────────────────
    with tab_delete:
        st.markdown("#### Remove a radar from the fleet")
        st.warning("⚠️ This removes the radar from AUDIT MATRIX. Audit history is preserved in HISTORY_LOG.")
        _del_bu_opts = ["All BUs"] + sorted(df_mat["BU"].dropna().unique().tolist()) if "BU" in df_mat.columns else ["All BUs"]
        del_bu_filt  = st.selectbox("Filter by BU", _del_bu_opts, key="del_bu_filt")
        _del_df      = df_mat if del_bu_filt=="All BUs" else df_mat[df_mat["BU"]==del_bu_filt]
        all_radars_del = sorted(_del_df["Radar Name"].dropna().astype(str).tolist())
        if not all_radars_del:
            st.info("No radars in this BU.")
        else:
            def _rlabel_d(r):
                rows=df_mat[df_mat["Radar Name"]==r]; t=str(rows["Type"].values[0] if len(rows)>0 else ""); s=str(rows["Site"].values[0] if len(rows)>0 else "")
                return " · ".join(p for p in [r,t,s] if p and p!="nan")
            sel_lbl_d   = st.selectbox("Select Radar to Remove", [_rlabel_d(r) for r in all_radars_del], key="del_sel")
            sel_r_d     = all_radars_del[[_rlabel_d(r) for r in all_radars_del].index(sel_lbl_d)]
            row_d       = df_mat[df_mat["Radar Name"]==sel_r_d].iloc[0]
            st.markdown(f"""<div style="border-left:4px solid #DC2626;border-radius:4px;padding:10px 14px;margin:8px 0;
                background:rgba(220,38,38,.08)"><b style="color:#DC2626">About to remove:</b><br>
                <b>{sel_r_d}</b> · {row_d.get("Type","")} · {row_d.get("Site","")} · {row_d.get("BU","")}</div>""",
                unsafe_allow_html=True)
            confirm_del = st.text_input("Type the radar name to confirm", placeholder=sel_r_d, key="del_confirm")
            if st.button("🗑️ Remove Radar", type="secondary", key="btn_delete"):
                if confirm_del.strip() != sel_r_d:
                    st.error(f"❌ Type exactly **{sel_r_d}** to confirm.")
                else:
                    try:
                        ws_am      = wb["AUDIT MATRIX"]
                        target_del = next((r[0].row for r in ws_am.iter_rows(min_row=2) if str(r[0].value)==sel_r_d), None)
                        if target_del:
                            ws_am.delete_rows(target_del); save_workbook(wb); load_wb_bytes.clear()
                            st.success(f"✅ **{sel_r_d}** removed."); st.rerun()
                        else:
                            st.error("Radar not found.")
                    except Exception as e:
                        st.error(f"❌ {e}")

    # ── TAB: IMPORT / BACKUP ──────────────────────────────────────────────────
    with tab_import:
        st.markdown("#### 📥 Import & Backup")
        imp_col, bak_col = st.columns([3,2])

        with bak_col:
            st.markdown("**💾 Backup current Master**")
            st.caption("Download before making changes.")
            try:
                with open(EXCEL_PATH,"rb") as _bf: _bk_bytes = _bf.read()
                st.download_button("⬇️ Download Backup", data=_bk_bytes,
                    file_name=f"Fleet_DQP_Master_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                st.caption(f"{len(df_mat)} radars · {len(_bk_bytes)//1024} KB")
            except Exception as _be:
                st.error(f"❌ {_be}")

        with imp_col:
            st.markdown("**📥 Import new Master**")
            st.caption("Only adds new radars (Radar Name / Site / Type / BU / Access). Audit history and scores from the uploaded file are **never imported** — your existing audit data is always preserved.")
            uploaded_master = st.file_uploader("Select Fleet_DQP_Master.xlsx", type=["xlsx"], key="master_upload")
            if uploaded_master is not None:
                try:
                    import io as _io
                    _raw  = uploaded_master.read()
                    _wb_n = load_workbook(_io.BytesIO(_raw))
                    if "AUDIT MATRIX" not in _wb_n.sheetnames:
                        st.error("❌ Missing AUDIT MATRIX sheet.")
                    else:
                        _ws_n    = _wb_n["AUDIT MATRIX"]
                        _hdrs_n  = [c.value for c in _ws_n[1]]
                        _missing_c = {"Radar Name","Site","Type","BU"} - set(_hdrs_n)
                        if _missing_c:
                            st.error(f"❌ Missing columns: {', '.join(_missing_c)}")
                        else:
                            _ra_idx = _hdrs_n.index("Remote Access") if "Remote Access" in _hdrs_n else None
                            _bu_idx = _hdrs_n.index("BU")
                            _uploaded = []
                            for _r in range(2, _ws_n.max_row+1):
                                _rn = _ws_n.cell(_r,1).value
                                if not _rn: continue
                                _ra = str(_ws_n.cell(_r,_ra_idx+1).value or "").strip() if _ra_idx is not None else ""
                                _bu = str(_ws_n.cell(_r,_bu_idx+1).value or "").strip()
                                _uploaded.append({"row":_r,"Radar Name":str(_rn).strip(),
                                    "Site":str(_ws_n.cell(_r,_hdrs_n.index("Site")+1).value or "").strip(),
                                    "Type":str(_ws_n.cell(_r,_hdrs_n.index("Type")+1).value or "").strip(),
                                    "BU":_bu,"Remote Access":_ra})
                            _existing  = set(df_mat["Radar Name"].astype(str).str.strip())
                            _new_ones  = [r for r in _uploaded if r["Radar Name"] not in _existing]
                            _dupes     = [r for r in _uploaded if r["Radar Name"] in _existing]
                            _ACCESS    = ["GSS","VPN","Bomgar","TeamViewer","Anydesk","Customer"]
                            _miss_ra   = [r for r in _new_ones if r["Remote Access"] not in _ACCESS]
                            s1,s2,s3   = st.columns(3)
                            s1.metric("New radars",    len(_new_ones), help="Will be added")
                            s2.metric("Already exist", len(_dupes),    help="Skipped — data preserved")
                            s3.metric("Missing Access",len(_miss_ra),  help="Need Remote Access")
                            if _miss_ra:
                                st.warning(f"⚠️ {len(_miss_ra)} radar(s) need Remote Access.")
                                if "import_ra_edits" not in st.session_state:
                                    st.session_state["import_ra_edits"] = {}
                                for _bu_name, _bu_rows in sorted({r["BU"]:[rr for rr in _miss_ra if rr["BU"]==r["BU"]] for r in _miss_ra}.items()):
                                    with st.expander(f"🌎 {_bu_name} — {len(_bu_rows)} radars", expanded=True):
                                        _ch = st.columns([2,2,1,2])
                                        _ch[0].markdown("**Radar**"); _ch[1].markdown("**Site**")
                                        _ch[2].markdown("**Type**");  _ch[3].markdown("**Remote Access**")
                                        for _ri in _bu_rows:
                                            _key = f"ra_{_ri['Radar Name']}"
                                            _c0,_c1,_c2,_c3 = st.columns([2,2,1,2])
                                            _c0.markdown(f"`{_ri['Radar Name']}`"); _c1.markdown(_ri["Site"]); _c2.markdown(f"`{_ri['Type']}`")
                                            _cur = st.session_state["import_ra_edits"].get(_key,"")
                                            _sel = _c3.selectbox("",[""] + _ACCESS,
                                                index=([""] + _ACCESS).index(_cur) if _cur in _ACCESS else 0,
                                                key=_key, label_visibility="collapsed")
                                            st.session_state["import_ra_edits"][_key] = _sel
                                _still_empty = [r for r in _miss_ra if not st.session_state["import_ra_edits"].get(f"ra_{r['Radar Name']}","")]
                            else:
                                _still_empty = []
                                if "import_ra_edits" in st.session_state: del st.session_state["import_ra_edits"]
                            if _still_empty:
                                st.error(f"❌ {len(_still_empty)} radar(s) still need Remote Access.")
                            elif _new_ones:
                                st.divider()
                                if st.button("➕ Add New Radars to Fleet", type="primary", key="btn_merge_master"):
                                    try:
                                        ws_am   = wb["AUDIT MATRIX"]
                                        headers = [c.value for c in ws_am[1]]
                                        col_map = {h:i+1 for i,h in enumerate(headers) if h}
                                        bu_col  = ensure_bu_column(ws_am)
                                        for _nr in _new_ones:
                                            nr  = ws_am.max_row+1
                                            ws_am.cell(nr,col_map.get("Radar Name",1)).value = _nr["Radar Name"]
                                            ws_am.cell(nr,col_map.get("Site",2)).value       = _nr["Site"]
                                            ws_am.cell(nr,col_map.get("Type",3)).value       = _nr["Type"]
                                            ws_am.cell(nr,bu_col).value                      = _nr["BU"]
                                            _ra_f = st.session_state.get("import_ra_edits",{}).get(f"ra_{_nr['Radar Name']}", _nr["Remote Access"])
                                            if "Remote Access" in col_map: ws_am.cell(nr,col_map["Remote Access"]).value = _ra_f
                                            if "PA"            in col_map: ws_am.cell(nr,col_map["PA"]).value            = "N/A" if _nr["Type"] in ("XT","SAR-X") else "Yes"
                                        save_workbook(wb); load_wb_bytes.clear()
                                        if "import_ra_edits" in st.session_state: del st.session_state["import_ra_edits"]
                                        st.success(f"✅ {len(_new_ones)} radars added. Existing data preserved."); st.rerun()
                                    except Exception as _e:
                                        st.error(f"❌ {_e}")
                            else:
                                st.info("✅ No new radars to add — all already exist.")
                except Exception as e:
                    st.error(f"❌ {e}")
            st.divider()
            st.caption("💡 Always download a backup before importing.")

    # ── TAB: ADMIN ────────────────────────────────────────────────────────────
    with tab_admin:
        st.markdown("#### 🔧 Data Administration")
        st.caption("Manage audit history — delete specific records or clean seed data.")

        # Reload wb/df for admin context
        wb     = get_workbook()
        df_mat = get_audit_matrix(wb)

        admin_tab1, admin_tab2, admin_tab3 = st.tabs([
            "🗑️ Delete by Radar", "🏭 Delete by Site", "💣 Reset / Clean"
        ])

        with admin_tab1:
            st.caption("Delete all audit history for one radar — keeps the radar in the fleet.")
            df_hist_adm = get_history(wb)
            if df_hist_adm.empty:
                st.info("No audit history yet.")
            else:
                audited_radars = sorted(df_hist_adm["Radar Name"].dropna().unique().tolist())
                del_radar = st.selectbox("Select radar to clear", audited_radars, key="admin_del_radar")
                n_entries = len(df_hist_adm[df_hist_adm["Radar Name"] == del_radar])
                last_score = df_hist_adm[df_hist_adm["Radar Name"] == del_radar]["Score Captured"].iloc[-1] if n_entries else None
                st.markdown(f"""<div style="background:#FEF2F2;border-radius:6px;padding:8px 12px;
                    border-left:3px solid #DC2626;font-size:.85rem">
                    <b>{del_radar}</b> — {n_entries} record(s) will be deleted · Last score: {f"{last_score:.0%}" if pd.notna(last_score) else "—"}
                </div>""", unsafe_allow_html=True)
                if st.button(f"🗑️ Clear audits for {del_radar}", key="btn_del_radar", type="secondary"):
                    try:
                        wb_d = get_workbook()
                        ws_hl = wb_d["HISTORY_LOG"]
                        rows_to_delete = [r for r in ws_hl.iter_rows(min_row=2) if r[0].value == del_radar]
                        for row in rows_to_delete:
                            for c in row: c.value = None
                        all_rows = [list(r) for r in ws_hl.iter_rows(min_row=2, values_only=True)]
                        non_empty = [r for r in all_rows if any(v is not None for v in r)]
                        for i, row in enumerate(ws_hl.iter_rows(min_row=2)):
                            if i < len(non_empty):
                                for j, c in enumerate(row): c.value = non_empty[i][j]
                            else:
                                for c in row: c.value = None
                        ws_am2 = wb_d["AUDIT MATRIX"]
                        hdrs2 = [c.value for c in ws_am2[1]]
                        META_R2 = {"Radar Name","Site","Type","Remote Access","PA","BU","Last Audit","Total Audits","Auditor"}
                        for row in ws_am2.iter_rows(min_row=2):
                            if str(row[0].value) == del_radar:
                                for c in row:
                                    h = hdrs2[c.column-1] if c.column <= len(hdrs2) else ""
                                    if h not in META_R2: c.value = None
                        save_workbook(wb_d); load_wb_bytes.clear()
                        st.success(f"✅ Audit data cleared for {del_radar}."); st.rerun()
                    except Exception as e:
                        st.error(f"❌ {e}")

        with admin_tab2:
            st.caption("Delete all audit history for every radar at one site.")
            df_hist_adm2 = get_history(wb)
            if df_hist_adm2.empty:
                st.info("No audit history yet.")
            else:
                audited_sites = sorted(df_hist_adm2["Site"].dropna().unique().tolist())
                del_site = st.selectbox("Select site to clear", audited_sites, key="admin_del_site")
                site_entries = df_hist_adm2[df_hist_adm2["Site"] == del_site]
                n_site_radars = site_entries["Radar Name"].nunique()
                n_site_entries = len(site_entries)
                st.markdown(f"""<div style="background:#FEF2F2;border-radius:6px;padding:8px 12px;
                    border-left:3px solid #DC2626;font-size:.85rem">
                    <b>{del_site}</b> — {n_site_entries} record(s) across {n_site_radars} radar(s) will be deleted
                </div>""", unsafe_allow_html=True)
                confirm_site = st.text_input("Type the site name to confirm", placeholder=del_site, key="admin_confirm_site")
                if st.button(f"🗑️ Clear all audits for {del_site}", key="btn_del_site", type="secondary"):
                    if confirm_site.strip() != del_site:
                        st.error(f"Type exactly: **{del_site}**")
                    else:
                        try:
                            wb_d2 = get_workbook()
                            site_radars_list = site_entries["Radar Name"].unique().tolist()
                            ws_hl2 = wb_d2["HISTORY_LOG"]
                            all_rows2 = [list(r) for r in ws_hl2.iter_rows(min_row=2, values_only=True)]
                            non_empty2 = [r for r in all_rows2 if any(v is not None for v in r) and r[1] != del_site]
                            for i, row in enumerate(ws_hl2.iter_rows(min_row=2)):
                                if i < len(non_empty2):
                                    for j, c in enumerate(row): c.value = non_empty2[i][j]
                                else:
                                    for c in row: c.value = None
                            ws_am3 = wb_d2["AUDIT MATRIX"]
                            hdrs3 = [c.value for c in ws_am3[1]]
                            META_R3 = {"Radar Name","Site","Type","Remote Access","PA","BU","Last Audit","Total Audits","Auditor"}
                            for row in ws_am3.iter_rows(min_row=2):
                                if str(row[0].value) in site_radars_list:
                                    for c in row:
                                        h = hdrs3[c.column-1] if c.column <= len(hdrs3) else ""
                                        if h not in META_R3: c.value = None
                            save_workbook(wb_d2); load_wb_bytes.clear()
                            st.success(f"✅ All audit data cleared for {del_site}."); st.rerun()
                        except Exception as e:
                            st.error(f"❌ {e}")

        with admin_tab3:
            st.markdown("**🧹 Fix phantom 'Up to Date' radars**")
            st.info(
                "If radars appear as 'Up to Date' but you haven't audited them, "
                "it means the AUDIT MATRIX has old demo/seed dates in the Last Audit column. "
                "Click below to clear those fields — **does not affect HISTORY_LOG or real audit scores.**"
            )
            if st.button("🧹 Clear Last Audit dates from AUDIT MATRIX", key="btn_clear_matrix_dates", type="primary"):
                try:
                    wb_cm = get_workbook()
                    ws_cm = wb_cm["AUDIT MATRIX"]
                    hdrs_cm = [c.value for c in ws_cm[1]]
                    AUDIT_FIELDS = {"Last Audit","Total Audits","Auditor","SCORE","ISSUES DETECTED",
                                    "RECOMMENDED ACTION","HELPER_SCORE","HELPER_DATE","HELPER_STATUS"}
                    cleared = 0
                    for row in ws_cm.iter_rows(min_row=2):
                        if not row[0].value: continue
                        for cell in row:
                            h = hdrs_cm[cell.column-1] if cell.column <= len(hdrs_cm) else ""
                            if h in AUDIT_FIELDS and cell.value is not None:
                                cell.value = None
                                cleared += 1
                    save_workbook(wb_cm); load_wb_bytes.clear()
                    st.success(f"✅ Cleared {cleared} audit date fields from AUDIT MATRIX. Expirations will now show correctly."); st.rerun()
                except Exception as _e:
                    st.error(f"❌ {_e}")

            st.divider()
            st.markdown("**🧹 Clear seed/demo data from HISTORY_LOG**")
            st.info("Removes HISTORY_LOG entries with no Score — clears fake seed data without touching real audits.")
            if st.button("🧹 Clear HISTORY_LOG seed data (no score)", key="btn_clear_seed"):
                try:
                    wb_cs = get_workbook()
                    ws_cs = wb_cs["HISTORY_LOG"]
                    rows_del = [row[0].row for row in ws_cs.iter_rows(min_row=2)
                                if row[0].value and (row[4].value is None or str(row[4].value).strip() in ["","nan","None"])]
                    for r in sorted(rows_del, reverse=True):
                        ws_cs.delete_rows(r)
                    save_workbook(wb_cs); load_wb_bytes.clear()
                    st.success(f"✅ Removed {len(rows_del)} seed entries from HISTORY_LOG."); st.rerun()
                except Exception as _e:
                    st.error(f"❌ {_e}")
            st.divider()
            st.markdown("**💣 Reset ALL audit data**")
            st.warning("⚠️ Permanently deletes ALL audit history and DQP answers for the entire fleet.")
            confirm = st.text_input("Type RESET to confirm", key="reset_confirm")
            if st.button("💣 Clear All Audit Data", type="secondary", key="btn_reset_all"):
                if confirm == "RESET":
                    try:
                        wb_r = get_workbook()
                        ws_h = wb_r["HISTORY_LOG"]
                        for row in ws_h.iter_rows(min_row=2):
                            for c in row: c.value = None
                        ws_a = wb_r["AUDIT MATRIX"]
                        hdrs = [c.value for c in ws_a[1]]
                        META_R = {"Radar Name","Site","Type","Remote Access","PA","BU"}
                        for row in ws_a.iter_rows(min_row=2):
                            for c in row:
                                if hdrs[c.column-1] not in META_R: c.value = None
                        save_workbook(wb_r); load_wb_bytes.clear()
                        st.success("✅ All audit data cleared."); st.rerun()
                    except Exception as e:
                        st.error(f"❌ {e}")
                else:
                    st.error("Type RESET exactly to confirm.")

    st.divider()
    _n_bus = df_mat["BU"].nunique() if "BU" in df_mat.columns else "—"
    _fc1, _fc2 = st.columns([3,1])
    _fc1.markdown(f"**Fleet: {len(df_mat)} radars · {df_mat['Site'].nunique()} sites · {_n_bus} BUs**")
    # CSV export of full fleet
    _export_cols = [c for c in ["Radar Name","Site","BU","Type","Remote Access","Last Audit","Total Audits"] if c in df_mat.columns]
    _csv = df_mat[_export_cols].to_csv(index=False).encode("utf-8")
    _fc2.download_button("⬇️ Export CSV", data=_csv,
        file_name=f"Fleet_DQP_{datetime.now().strftime('%Y%m%d')}.csv",
        mime="text/csv", use_container_width=True)
